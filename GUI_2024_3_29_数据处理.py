import tkinter as tk
from tkinter import ttk
from tkinter import filedialog, messagebox
import pandas as pd
import copy
import openpyxl
from openpyxl.styles import Font, PatternFill, Color
from datetime import datetime
import numpy as np
import os
import time
from threading import Thread
import math


# import queue


# 创建GUI图形界面
class DataProcessingApp:
    start_time = None
    second_start_time = None

    def __init__(self, root):
        self.root = root
        self.root.title("老王出品 ： 远程数据处理  7.0 V")

        self.file_path = None  # 初始化 file_path 属性为 None
        self.save_path = None  # 初始化 save_path 属性为 None

        # 创建按钮和标签等控件
        self.select_file_button = tk.Button(self.root, text="选择文件", command=self.select_file)
        # 布局控件,选择文件 select_file_button
        self.select_file_button.pack()
        # 创建一个Entry用于显示文件名
        self.select_name_entry = tk.Entry(self.root, width=60)
        self.select_name_entry.pack()

        self.save_file_button = tk.Button(self.root, text="选择保存位置", command=self.save_file)
        self.save_file_button.pack()
        self.save_file_entry = tk.Entry(self.root, width=60)  # 用于显示保存路径
        self.save_file_entry.pack()

        # 创建了一个Frame并将其放置在主窗口中。然后，我们将voltage_label、min_voltage_entry和
        # max_voltage_entry小部件分别放置在这个Frame中，并使用side选项将它们水平排列在同一行上。
        voltage_frame = tk.Frame(self.root)
        voltage_frame.pack()

        voltage_label = tk.Label(voltage_frame, text="电堆电压选择范围：")
        voltage_label.pack(side=tk.LEFT)
        self.mix_voltage_entry = tk.Entry(voltage_frame, width=10)
        self.mix_voltage_entry.insert(0, "92")
        self.mix_voltage_entry.pack(side=tk.LEFT)
        self.max_voltage_entry = tk.Entry(voltage_frame, width=10)
        self.max_voltage_entry.insert(0, "125")
        self.max_voltage_entry.pack(side=tk.LEFT)
        voltage_label1 = tk.Label(voltage_frame, text="(单位：V)")
        voltage_label1.pack(side=tk.RIGHT)

        # 创建了一个Frame并将其放置在主窗口中。然后，我们将voltage_label、min_voltage_entry和
        # max_voltage_entry小部件分别放置在这个Frame中，并使用side选项将它们水平排列在同一行上。
        voltage_frame1 = tk.Frame(self.root)
        voltage_frame1.pack(pady=(10))

        self.process_button = tk.Button(voltage_frame1, text="发电数据处理",
                                        command=self.background_process_data,
                                        width=15)
        self.process_button.pack(side=tk.RIGHT, padx=(50, 5))
        self.process_button_disabled = True  # 用于跟踪按钮状态的标志

        # 合并表格里面的多余的页为一页
        self.excel_process_button = tk.Button(voltage_frame1, text="表格页合并",
                                              command=self.background_excel_process_data,
                                              width=15)
        self.excel_process_button.pack(side=tk.RIGHT)
        self.excel_process_button_disabled = True  # 用于跟踪按钮状态的标志

        # 处理待机当天燃料的消耗（当天不发电才会触发）
        self.no_process_button = tk.Button(voltage_frame1, text="待机燃料计算",
                                           command=self.background_no_process_data,
                                           width=15)
        self.no_process_button.pack(side=tk.LEFT, padx=(5, 50))
        self.no_process_button_disabled = True  # 用于跟踪按钮状态的标志

        #
        instruction_button = tk.Button(self.root, text="使用说明", command=self.show_instruction)
        instruction_button.pack()

        # 创建进度条
        self.progress = ttk.Progressbar(root, orient="horizontal", length=200, mode="determinate")
        # 设置进度条的总值为100
        self.progress['maximum'] = 205
        self.progress.pack(pady=15)

        # 创建队列用于线程间通信
        # self.queue = queue.Queue()

    def select_file(self):
        # 使用文件对话框获取文件路径
        self.file_path = filedialog.askopenfilename()
        self.select_name_entry.delete(0, 'end')  # 清空文本框
        self.select_name_entry.insert(0, self.file_path)  # 将选择的文件名插入文本框

    def save_file(self):
        save_file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if save_file_path:
            self.save_path = os.path.join(save_file_path)
            self.save_file_entry.delete(0, "end")
            self.save_file_entry.insert(0, self.save_path)  # 将完整的文件路径赋值给相应的变量或更新界面上的显示

    def show_instruction(self):
        instruction_window = tk.Toplevel(self.root)
        instruction_window.title("操作说明")

        instruction_text = tk.Text(instruction_window, height=30, width=50)
        instruction_text.insert(tk.END, "作者：老王\n\n"
                                        "出品原由： 世界因懒人而进步 ！！ ！ \n\n"
                                        '***************************************\n\n'
                                        "数值必须为网页远程下载的原始数据\n\n"
                                        "本程序只针对COWIN网页远程下载的数据进行筛选 !!!\n\n"
                                        "“电堆电压选择范围”为电堆电压筛选\n"
                                        "计算电压出来的电压值为筛选范围的平均电压值\n"
                                        "众宇电堆电堆电压值默认范围：92-125\n"
                                        "攀业电堆电压值范围：65-125\n"
                                        "该值可以根据自己需要进行调整\n"
                                        '\n\n（Stapow）A/B（FcB_StackP）电堆功率计算=（前10个最大值）/10'
                                        '\n如果不满足10个值，则有多少值计数多少个值的平均值（0不参与计算）'
                                        '\n\n总功率计算=[（A+B功率）前10个最大值 ] / 10'
                                        '\n如果不满足10个值，则有多少值计数多少个值的平均值（0不参与计算）'
                                        '\n\n（S_RemFuelIn）燃料消耗值计算，为小水箱里面液位每次减少的值求总的和，如果计算出来的燃烧消耗值为0，则默认给0.3'
                                        "\n\n版本更新 2.0 V ：\n"
                                        "excel 表格读取内容为 all sheet "
                                        '\n\n版本更新 3.0 V 2024.1.16 ：\n'
                                        '针对（FcB_StackP）列表数据为null情况做出筛选，去掉null部分，不参与计算（如：楼下机房，白石等老式发电机远程数据）'
                                        '\n\n版本更新 4.0 V 2024.1.27 ：\n'
                                        '（S_CurVol）新增母线电压：值为母线电压的平均值'
                                        '\n\n版本更新 5.0 V 2024.3.20 ：\n'
                                        '新增表格页面（sheet）合并按键：在有多个页面的excel表格中，将所有页面（all sheet）合并成一个页面（sheet），并保存'
                                        '\n\n版本更新 6.0 V 2024.3.21 ：\n'
                                        '新增进度条：实时显示程序运行进度\n'
                                        '新增待机燃料消耗计算和产氢次数计数：如果当天没有发电（一次都没有），则会计算当天待机时燃料的消耗和产氢次数'
                                        '\n\n版本更新 6.5 V 2024.3.22 ：\n'
                                        '新增警告：如果读取文件格式不是".xlsx“会弹出警告框，或者文件有上锁/加密，都会弹出警告框，中止程序运行\n'
                                        '\n\n版本更新 7.0 V 2024.3.29 ：\n'
                                        '发电数据统计：新增外置/内置燃料的毫米（mm）值计算。如果没有升(L)，改用毫米（mm）计算燃料消耗。如：白石，楼下机房。燃料消耗率(L.kWh -1)不参与计算，值为0. \n'
                                        '待机燃料计算：'
                                        '\n 1.新增平均产氢时间，值为每次产氢的间隔时间平均值'
                                        '\n 2.新增‘待机条件’判断：如果电堆电压‘StaV’全部为0，或者电堆功率‘Stapow’全部为0，则为待机待机状态，没有发电。（原本条件：整机开关’MSw‘全部都是False，则为待机状态）'
                                        '\n 3.新增备注条件：如果数据量（总行数）小于3500，则给备注加上注释。数据量（总行数）小于多少。因为数据太少，算出来的值不准确'

                                )
        instruction_text.config(state=tk.DISABLED)
        instruction_text.pack(fill=tk.BOTH, expand=True)

    def show_save_success_message(self, save_path):
        message = "文件保存成功！\n保存路径为：{}".format(save_path)
        messagebox.showinfo("保存成功", message)

    def show_save_fail_message(self, save_path):
        message = "文件保存失败！\n读取路径为：{}\n显示没有发电  ！  ！  ！".format(save_path)
        messagebox.showinfo("保存失败", message)

    def show_save_fail_message_1(self, save_path):
        message = "文件保存失败！\n读取路径为：{}\n显示没有发电  ！  ！  ！".format(save_path)
        messagebox.showinfo("保存失败", message)

    def show_read_error(self, file_path):
        message = f" 读取 '.xlsx' 文件失败 ！\n 请检查文件格式是否为'.xlsx'格式 ,文件是否有加密/上锁 ！！！\n 文件路径：{file_path}\n  "
        messagebox.showinfo('读取失败', message)

    ##############    开启线程    ##################

    def background_excel_process_data(self):

        # 当按下“合并数据”时，禁用“发电数据处理”按钮
        self.process_button.config(state=tk.DISABLED)
        self.process_button_disabled = False

        # 当按下“合并数据”时，禁用“发电数据处理”按钮
        self.no_process_button.config(state=tk.DISABLED)
        self.no_process_button_disabled = False

        threa = Thread(target=self.excel_process_data)
        threa.start()

    # 当按下“合并数据”时，禁用“发电数据处理”按钮
    def excel_process_data(self):

        self.progress.start()
        self.progress['value'] += 0  # 逐步增加进度条值
        df = []

        adress1 = self.file_path  # 读取文件路径。将选择的文件路径赋值给adress1变量
        adress3 = self.save_path  # 保存文件路径

        if not adress1 or not adress3:  # 假设 self.file_path 和 self.save_path 分别表示文件路径和保存路径

            messagebox.showerror("错误", "请选择文件路径和文件保存路径")

            for _ in range(105):
                self.progress['value'] += 1  # 如果，异常。满值进度条值
                self.progress.update()
                time.sleep(0.0001)  # 微小的延迟，实现平滑更新

            # 表格页合并完成后，重新启用“发电数据处理”按钮 。条件为假时，执行下面代码
            if not self.process_button_disabled:
                self.process_button.config(state=tk.NORMAL)

            # 表格页合并完成后，重新启用“发电数据处理”按钮 。条件为假时，执行下面代码
            if not self.no_process_button_disabled:
                self.no_process_button.config(state=tk.NORMAL)

            # 停止进度条动画
            self.progress.stop()
        try:
            if os.path.exists(adress1):  # 检查文件（文件名，文件路径是对得上）是否存在，不存在则结束程序
                try:

                    xl = pd.ExcelFile(adress1)  # 使用 pd.ExcelFile() 方法打开 Excel 文件

                    self.progress['value'] += 1  # 逐步增加进度条值
                    self.progress.update()
                    time.sleep(0.05)  # 微小的延迟，实现平滑更新

                    # df = pd.DataFrame()  # 创建一个空的数据框
                    for sheet_name in xl.sheet_names:  # 遍历文件中的所有 sheet
                        one_sheet = xl.parse(sheet_name)  # 读取当前 sheet 的数据
                        df.append(one_sheet)  # 将读取的数据合并到 all_data 中

                    # 第一阶段：读取和合并数据完成后更新进度条
                    for _ in range(10):
                        self.progress['value'] += 1  # 逐步增加进度条值
                        self.progress.update()
                        time.sleep(0.05)  # 微小的延迟，实现平滑更新

                    # 使用 pd.concat() 方法将所有数据框连接成一个
                    df = pd.concat(df, ignore_index=True)

                    # 第一阶段：读取和合并数据完成后更新进度条
                    for _ in range(20):
                        self.progress['value'] += 1  # 逐步增加进度条值
                        self.progress.update()
                        time.sleep(0.05)  # 微小的延迟，实现平滑更新

                except FileNotFoundError:
                    print(f"文件 {adress1} 不存在，已跳过")

                df.to_excel(adress3, index=False, engine='openpyxl')

                # 第二阶段：数据写入 Excel 文件完成后更新进度条
                for _ in range(20):
                    self.progress['value'] += 1  # 逐步增加进度条值
                    self.progress.update()
                    time.sleep(0.05)  # 微小的延迟，实现平滑更新

                # 打开现有的Excel文件
                workbook = openpyxl.load_workbook(adress3)
                # 选择第一个工作表
                sheet = workbook.active

                def excel_width_height():
                    # 设置第一行的行高
                    sheet.row_dimensions[1].height = 50
                    # 设置第一列和第二列的宽度为 25
                    sheet.column_dimensions['A'].width = 21  # 第一列
                    sheet.column_dimensions['B'].width = 21  # 第二列
                    # 设置其余列的宽度为 10
                    for col in sheet.columns:
                        if col[0].column_letter not in ['A']:
                            sheet.column_dimensions[col[0].column_letter].width = 15

                thread1 = Thread(target=excel_width_height)
                thread1.start()

                def excel_task():
                    # 遍历第一行的所有单元格，并为每个单元格对象同时设置自动换行、水平居中和垂直居中。
                    for cell in sheet[1]:
                        cell.alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal='center',
                                                                   vertical='center')
                        # 创建Font对象并设置字体大小,字体大小为14，加粗
                        cell.font = Font(size=16, bold=True)
                        # 创建PatternFill对象并设置RGB(178,180,164)背景
                        # 对应的十六进制颜色代码是'B2B4A4'
                        cell.fill = PatternFill(start_color='B2B4A4', end_color='B2B4A4', fill_type='solid')

                thread2 = Thread(target=excel_task)
                thread2.start()

                # 最后阶段：设置单元格格式完成后更新进度条
                for _ in range(20):
                    self.progress['value'] += 1  # 逐步增加进度条值
                    self.progress.update()
                    time.sleep(0.05)  # 微小的延迟，实现平滑更新

                # 等待线程执行完成后，才进入下下一步
                thread2.join()
                thread1.join()

                workbook.save(adress3)

                # 完成所有操作后更新进度条到100%
                for _ in range(10):
                    self.progress['value'] += 1  # 逐步增加进度条值
                    self.progress.update()
                    time.sleep(0.05)  # 微小的延迟，实现平滑更新

                # 表格页合并完成后，重新启用“发电数据处理”按钮 。条件为假时，执行下面代码
                if not self.process_button_disabled:
                    self.process_button.config(state=tk.NORMAL)

                # 表格页合并完成后，重新启用“发电数据处理”按钮 。条件为假时，执行下面代码
                if not self.no_process_button_disabled:
                    self.no_process_button.config(state=tk.NORMAL)

                # 停止进度条动画
                self.progress.stop()
                print(f"\n文件保存成功 ！! ! ")
                print(f"文件保存路径 ：{adress3}")
                self.show_save_success_message(adress3)
            else:
                print(f"文件 {adress1} 不存在，已跳过")

                # self.show_save_fail_message(adress1)
        except ValueError:

            self.show_read_error(adress1)

            for _ in range(105):
                self.progress['value'] += 1  # 如果，异常。满值进度条值
                self.progress.update()
                time.sleep(0.0001)  # 微小的延迟，实现平滑更新

                # 表格页合并完成后，重新启用“发电数据处理”按钮 。条件为假时，执行下面代码
            if not self.process_button_disabled:
                self.process_button.config(state=tk.NORMAL)

                # 表格页合并完成后，重新启用“发电数据处理”按钮 。条件为假时，执行下面代码
            if not self.no_process_button_disabled:
                self.no_process_button.config(state=tk.NORMAL)

                # 停止进度条动画
            self.progress.stop()

    ##############    关闭线程    ##################

    # ==========================================================#

    ##############    开启线程    ##################

    # 发电数据处理

    # 开启多线程主，调用函数，调用process_data
    def background_process_data(self):

        # 当按下“发电数据”时，禁用“合并数据”按钮
        self.excel_process_button.config(state=tk.DISABLED)
        self.excel_process_button_disabled = False
        # 当按下“合并数据”时，禁用“发电数据处理”按钮
        self.no_process_button.config(state=tk.DISABLED)
        self.no_process_button_disabled = False

        threa = Thread(target=self.process_data)
        threa.start()

    ##############    关闭线程    ##################

    def process_data(self):

        self.progress.start()
        self.progress['value'] += 0  # 逐步增加进度条值

        # 打印行号和列的数据
        A_Power_values = []
        B_Power_values = []
        power_values = []  # 储存发电时的功率值
        IC_value = []  # 储存发电时的芯片温度值
        Topgen_value = []  # 储存每次发电，开始/结束的发电量值
        Once_Topgen_value = []  # 储存，每次发电的发电量。用于算出总发电量
        Time_value = []  # 储存每次发电，开始/结束时间的值
        Time_diffs = []  # 储存，每次发电的时间的时长。用于算出总发电时间
        differences = []
        total_sum = 0
        fuel_levels = []
        last_fuel_levels = []
        S_RemFuelIn_value = []
        positive_differences = []
        calculate_positive_differences = []
        Once_S_RemFuelIn = []
        B_StackV_value = []
        A_StackV_value = []
        B_List = []
        A_List = []
        last_A_List = []
        last_B_List = []
        HGretem_value = []  # 发电时，储存 重整室温度的值到列表 HGretem_value
        Hfetem_value = []  # 发电时，储存 重整室温度的值到列表 Hfetem_value
        HGretem_list = []
        Hfetem_list = []
        last_HGretem_list = []
        last_Hfetem_list = []
        start_datatime = []
        end_datatime = []
        start_S_RemFuelIn = []
        end_S_RemFuelIn = []
        start_Topgen = []
        end_Topgen = []
        start_S_RemFuelOut = []
        end_S_RemFuelOut = []
        Stwtims = []
        Fuel_consumption = None
        current_voltage = []
        current_voltage_value = []
        start_current_voltage = []
        end_current_voltage = []
        everytime_current_voltage = []
        current_voltage_List_value = []
        last_current_voltage_List_value = []

        everytime_Topgen = []
        everytime_A_power = []
        everytime_B_power = []
        everytime_power = []
        everytime_IC = []
        everytime_A_StackV = []
        everytime_B_StackV = []
        everytime_max_HGretem = []
        everytime_min_HGretem = []
        everytime_max_Hfetem = []
        everytime_min_Hfetem = []
        everytime_Fuel_consumption = []
        df = []
        system_state = []

        copy_everytime_A_StackV = []
        copy_everytime_B_StackV = []
        copys_everytime_A_StackV = []
        copys_everytime_B_StackV = []
        copysS_everytime_A_StackV = []
        copysS_everytime_B_StackV = []
        modified_A_StackV = []
        modified_B_StackV = []
        count_end_datatime = []
        fuel_List_value = []
        last_A_power_value_list = []
        last_B_power_value_list = []
        last_power_value_list = []
        power_list = []
        start_time = None
        second_start_time = None
        copy_start_datatime = []
        copy_end_datatime = []
        count_datatime = []  # 开始时间+结束时间，放入一个列表里面。除以2余0.证明当天发电，开始和结束成一对。用于计算当天没有结束时的判断
        first_start_datatime = 0
        second_end_datatime = 0

        df_list = []

        true_LiqlelL = []  # 外置液位mm
        true_LiqlelM = []  # 内置液位mm

        start_LiqlelL = []
        end_LiqlelL = []

        start_LiqlelM = []
        end_LiqlelM = []

        # 使用self.mix_voltage_entry.get()
        # 来获取self.mix_voltage_entry中的数据，并将其存储在mix_voltage_data变量中
        mix_voltage_data = float(self.mix_voltage_entry.get())
        max_voltage_data = float(self.max_voltage_entry.get())
        adress1 = self.file_path  # 读取文件路径。将选择的文件路径赋值给adress1变量
        adress3 = self.save_path  # 保存文件路径

        # print(f"\n {adress1} \n")
        # print( f"\n {adress3} \n")
        if not adress1 or not adress3:  # 假设 self.file_path 和 self.save_path 分别表示文件路径和保存路径
            messagebox.showerror("错误", "请选择文件路径和文件保存路径")

            for _ in range(105):
                self.progress['value'] += 1  # 如果，异常。满值进度条值
                self.progress.update()
                time.sleep(0.0001)  # 微小的延迟，实现平滑更新

            # 表格页合并完成后，重新启用“发电数据处理”按钮 。条件为假时，执行下面代码
            if not self.excel_process_button_disabled:
                self.excel_process_button.config(state=tk.NORMAL)

            # 处理待机当天燃料的消耗（，重新启用“处理待机当天燃料的消耗（”按钮 。条件为假时，执行下面代码
            if not self.no_process_button_disabled:
                self.no_process_button.config(state=tk.NORMAL)

            # 停止进度条动画
            self.progress.stop()
        try:
            if os.path.exists(adress1):  # 检查文件（文件名，文件路径是对得上）是否存在，不存在则结束程序
                try:
                    # os.chmod(adress3,0o755)
                    # 在这里进行对数据的处理和分析
                    # df = pd.read_excel(adress1)

                    # xl = pd.ExcelFile(adress1)  # 使用 pd.ExcelFile() 方法打开 Excel 文件
                    df_list = []  # 初始化一个空的DataFrame列表，用于存储每个工作表的数据
                    # 使用 'with' 语句打开Excel文件
                    with pd.ExcelFile(adress1) as xl:
                        for sheet_name in xl.sheet_names:  # 遍历文件中的所有 sheet
                            one_sheet = xl.parse(sheet_name)  # 读取当前 sheet 的数据
                            df_list.append(one_sheet)  # 将读取的数据添加到df_list中

                    df = pd.concat(df_list, ignore_index=True)  # 使用 pd.concat() 方法将所有数据框连接成一个

                    # 使用fillna()方法来替换DataFrame中的NaN值。如果你想要将所有的NaN值替换为0，可以直接调用方法 fillna(0)
                    df.fillna(0, inplace=True)

                    df['电堆总功率'] = df['Stapow'] + df['FcB_StackP']

                    # # notnull()方法来检查列中的每个元素是否为非空值，然后结合any()方法来判断是否有至少一个非空值存在。
                    # if df['FcB_StackP'].notnull().any():
                    #     df['电堆总功率'] = df['Stapow'] + df['FcB_StackP']
                    # else:
                    #     df['电堆总功率'] = df['Stapow']

                    # 选择要读取的列名
                    MSw = 'MSw'  # 开关状态
                    DateTime = 'DateTime'  # 时间
                    S_RemFuelIn = 'S_RemFuelIn'  # 内置水箱液位
                    S_RemFuelOut = 'S_RemFuelOut'  # 外置水箱液位
                    Topgen = 'Topgen'  # 发电量
                    IC_Temp = 'Chiptem'  # 芯片温度
                    A_Power = 'Stapow'  #
                    B_Power = 'FcB_StackP'
                    Power = '电堆总功率'
                    prev_row = None
                    B_StackV = 'FcB_StackV'  # 电堆B电压
                    A_StackV = 'StaV'  # 电堆A电压
                    HGretem = 'HGretem'  # 重整室温度
                    Hfetem = 'Hfetem'  # 提纯器温度
                    Stwtim = 'Stwtim'  # 发电次数
                    S_CurVol = 'S_CurVol'  # 母线电压

                    LiqlelL = 'LiqlelL'  # 外置液位（mm）
                    LiqlelM = 'LiqlelM'  # 内置液位（mm）

                    #   如果电压小于85，则跳过当天计算
                    if any(df['StaV'] > 60):
                        second_row = df.iloc[1]  # 这行代码将DataFrame中的第二行数据存储在变量second_row中，以便后续对第二行数据进行操作和分析
                        last_row = df.iloc[-1]  # 这行代码将DataFrame中的最后一行数据存储在变量last_row中，以便后续对最后一行数据进行操作和分析

                        # #  !!!  如果计算对象是 “众宇电堆” 筛选范围选择：  ９２ ＜＝ Ｘ ＜ １２５
                        # #  !!!  如果计算对象是 “攀业电堆” 筛选范围选择：  ７５ ＜＝ Ｘ ＜ １２０
                        # 对电堆电压算平均值 。
                        def calculate_filtered_average(data):
                            filtered_data = [x for x in data if mix_voltage_data <= x < max_voltage_data]  # 设置筛选范围
                            average = sum(filtered_data) / len(filtered_data) if len(filtered_data) > 0 else 0  # 计算平均值
                            return average

                        # 对发电功率算平均值,计算列表元素十个最大值平均值
                        def calculate_average(input_list):
                            # 去掉小于100的元素并重新生成列表
                            new_list = [x for x in input_list if x >= 100]

                            if len(new_list) > 10:  # 如果新列表元素个数大于10
                                top_values = sorted(new_list, reverse=True)[:10]  # 找出新列表元素十个最大值
                                average = sum(top_values) / 10  # 计算平均值
                                return average
                            elif len(set(new_list)) == 1:  # 如果所有元素都相等
                                return new_list[0]  # 返回任意一个元素的值作为平均值
                            elif 0 < len(new_list) <= 10:  # 如果新列表元素个数小于等于10且不为空
                                average = sum(new_list) / len(new_list)  # 计算所有元素的平均值
                                return average
                            else:
                                if len(new_list) == 0:  # 如果新列表为空
                                    return 0

                        print('\n ————————————————    一天计算开始    ————————————————    \n')

                        for index, row in df.iterrows():  # 这段代码会遍历 DataFrame df 中的每一行数据。

                            if prev_row is not None:  # 这段代码检查变量 prev_row 是否为非空值。

                                if row[MSw] == True:  # 如果MSW=TRUE，发电时，储存发电时间段内某列的数据
                                    A_Power_values.append(round(row[A_Power], 1))
                                    B_Power_values.append(round(row[B_Power], 1))
                                    power_values.append(round(row[Power], 1))  # 发电时，储存 功率 的值到列表 power_values
                                    IC_value.append(round(row[IC_Temp], 1))  # 发电时，储存 芯片温度 的值到列表 power_values

                                    S_RemFuelIn_value.append(
                                        round(row[S_RemFuelIn], 1))  # 发电时，储存 内置水箱剩余燃料 的值到列表 S_RemFuelIn_value
                                    B_StackV_value.append(round(row[B_StackV], 1))  # 发电时，储存 电堆B电压 的值到列表 B_StackV_value
                                    A_StackV_value.append(round(row[A_StackV], 1))  # 发电时，储存 电堆A电压 的值到列表 A_StackV_value
                                    HGretem_value.append(round(row[HGretem], 1))  # 发电时，储存 重整室温度的值到列表 HGretem_value
                                    Hfetem_value.append(round(row[Hfetem], 1))  # 发电时，储存 提纯室温度的值到列表 Hfetem_value
                                    current_voltage_value.append(
                                        round(row[S_CurVol], 1))  # 发电时，储存 母线电压的值到 current_voltage
                                    # print(f'上一个power_values功率值>>>>>>>>>>>>：{np.array(power_values) }')

                                    true_LiqlelM.append(
                                        round(row[LiqlelM], 2))  # 发电时，储存 内置水箱剩余燃料(mm) 的值到列表 true_LiqlelM
                                    true_LiqlelL.append(
                                        round(row[LiqlelL], 2))  # 发电时，储存 外置水箱剩余燃料(mm) 的值到列表 true_LiqlelL

                                if prev_row[MSw] == False and row[MSw] == True:  # 开始发电时间 。 如果MSW的上一个值=false,并且当前的值=true
                                    print(f"\n第一有开始 ###############\n")
                                    print(  # 在控制台上打印，显示
                                        f"开始发电时间：{row[DateTime]}      "
                                        f"内置水箱剩余燃料: {round(row[S_RemFuelIn], 1)}     "
                                        f"外置水箱剩余燃料: {round(row[S_RemFuelOut], 1)}    "

                                        f"内置水箱剩余燃料(mm): {round(row[LiqlelM], 1)} "
                                        f"外置水箱剩余燃料(mm): {round(row[LiqlelL], 1)} "

                                        f"总发电量:{round(row[Topgen], 1)}      ")
                                    Topgen_value.append(round(row[Topgen], 1))
                                    Time_value.append(row[DateTime])
                                    count_end_datatime.append(row[DateTime])
                                    second_start_time = row[DateTime]  # 用于后面当天发电缺少“开始发电”的判断
                                    # 创建列表用于储存输出到excel表格和数据
                                    start_datatime.append(row[DateTime])
                                    start_S_RemFuelIn.append(round(row[S_RemFuelIn], 1))
                                    start_Topgen.append(round(row[Topgen], 1))
                                    start_S_RemFuelOut.append(round(row[S_RemFuelOut], 1))

                                    start_LiqlelL.append(round(row[LiqlelL], 1))
                                    start_LiqlelM.append(round(row[LiqlelM], 1))

                                #  start_current_voltage.append(round(row[S_CurVol], 1))
                                else:

                                    if second_start_time is None and second_row[MSw] == True:  #
                                        print(f"\n第二没有开始 ************\n")
                                        print(
                                            f"开始发电时间：{second_row[DateTime]}      "
                                            f"内置水箱剩余燃料: {round(second_row[S_RemFuelIn], 1)}     "
                                            f"外置水箱剩余燃料: {round(second_row[S_RemFuelOut], 1)}"

                                            f"内置水箱剩余燃料(mm): {round(row[LiqlelM], 1)} "
                                            f"外置水箱剩余燃料(mm): {round(row[LiqlelL], 1)} "

                                            f"    总发电量:{round(second_row[Topgen], 1)}      ")
                                        Topgen_value.append(round(second_row[Topgen], 1))
                                        Time_value.append(second_row[DateTime])
                                        second_start_time = second_row[DateTime]
                                        count_end_datatime.append(second_row[DateTime])
                                        # 创建列表用于储存输出到excel表格和数据
                                        start_datatime.append(second_row[DateTime])
                                        copy_start_datatime.append(second_row[DateTime])
                                        first_start_datatime = len(copy_start_datatime)
                                        start_S_RemFuelIn.append(round(second_row[S_RemFuelIn], 1))
                                        start_Topgen.append(round(second_row[Topgen], 1))
                                        start_S_RemFuelOut.append(round(second_row[S_RemFuelOut], 1))

                                        start_LiqlelL.append(round(row[LiqlelL], 1))
                                        start_LiqlelM.append(round(row[LiqlelM], 1))

                                    # start_current_voltage.append(round(row[second_row[S_CurVol]], 1))
                                if prev_row[MSw] == True and row[MSw] == False:  # 结束发电时间。如果MSW的上一个值=true,并且当前的值=false
                                    print(
                                        f"结束发电时间：{prev_row[DateTime]}      "
                                        f"内置水箱剩余燃料: {round(prev_row[S_RemFuelIn], 1)}     "
                                        f"外置水箱剩余燃料: {round(prev_row[S_RemFuelOut], 1)}    "
                                        f"内置水箱剩余燃料(mm): {round(prev_row[LiqlelM], 1)} "
                                        f"外置水箱剩余燃料(mm): {round(prev_row[LiqlelL], 1)}"
                                        f"总发电量:{round(prev_row[Topgen], 1)}    ")

                                    print(len(count_end_datatime))  # 计算当天发电次数
                                    Topgen_value.append(round(prev_row[Topgen], 1))
                                    Time_value.append(prev_row[DateTime])
                                    start_time = prev_row[DateTime]  # 用于后面当天发电缺少“结束发电”的判断

                                    # 创建列表用于储存输出到excel表格和数据
                                    end_datatime.append(prev_row[DateTime])
                                    end_S_RemFuelIn.append(round(prev_row[S_RemFuelIn], 1))
                                    end_Topgen.append(round(prev_row[Topgen], 1))
                                    end_S_RemFuelOut.append(round(prev_row[S_RemFuelOut], 1))

                                    end_LiqlelL.append(round(prev_row[LiqlelL], 1))
                                    end_LiqlelM.append(round(prev_row[LiqlelM], 1))

                                    Once_Topgen = round(Topgen_value[-1] - Topgen_value[-2], 3)
                                    print(f"每次发电量(kw/h)：{Once_Topgen}")
                                    Once_Topgen_value.append(Once_Topgen)

                                    Stwtims.append(row[Stwtim])
                                    print(f"发电次数：{row[Stwtim]}")

                                    Time_diff = round(
                                        (pd.to_datetime(Time_value[-1]) - pd.to_datetime(
                                            Time_value[-2])).total_seconds() / 60,
                                        2)
                                    Time_diffs.append(Time_diff)
                                    print(f"每次发电时长(min)：{Time_diff}")

                                    mean_IC = round(sum(IC_value) / len(IC_value), 2)
                                    everytime_IC.append(mean_IC)
                                    print(f'芯片平均温度(℃):{mean_IC}')

                                    Once_RemFuelIn = 0

                                    # 一天只发一次电时，执行下面程序
                                    if len(count_end_datatime) == 1:
                                        current_voltage = round(sum(current_voltage_value) / len(current_voltage_value),
                                                                1)
                                        everytime_current_voltage.append(current_voltage)
                                        print(f'母线电压平均值(W)：{current_voltage}')
                                        # current_voltage.clear()

                                        calculate_A_power = round(calculate_average(A_Power_values), 1)
                                        everytime_A_power.append(calculate_A_power)
                                        A_Power_values.clear()
                                        print(f'A堆功率平均值(W)：{calculate_A_power}')

                                        calculate_B_power = round(calculate_average(B_Power_values), 1)
                                        everytime_B_power.append(calculate_B_power)
                                        B_Power_values.clear()
                                        print(f'B堆功率平均值(W)：{calculate_B_power}')

                                        calculate_power = round(calculate_average(power_values), 1)
                                        everytime_power.append(calculate_power)
                                        power_values.clear()
                                        print(f'总功率平均值(W)：{calculate_power}')

                                        print(f'S_RemFuelIn_value[0]：{S_RemFuelIn_value[0]}')
                                        if S_RemFuelIn_value[0] > 0:
                                            differences = [S_RemFuelIn_value[i] - S_RemFuelIn_value[i + 1] for i in
                                                           range(len(S_RemFuelIn_value) - 1)]
                                            positive_differences = [x for x in differences if x > 0]
                                            Once_RemFuelIn = round(sum(positive_differences), 2)
                                            if Once_RemFuelIn == 0:
                                                Once_RemFuelIn = 0.3
                                            Once_S_RemFuelIn.append(Once_RemFuelIn)
                                            print(f'每次发电消耗燃料（L）:{Once_RemFuelIn}')
                                            S_RemFuelIn_value.clear()  # 用完S_RemFuelIn_value列表后，要把列表清空，不然会叠加列表
                                        else:
                                            differences = round(start_LiqlelM[-1] - end_LiqlelM[-1], 1)
                                            if differences < 0:
                                                differences = 0
                                            Once_S_RemFuelIn.append(differences)
                                            print(f'每次发电消耗燃料（mm）:{differences}')
                                            # print(f'液位(mm)******** ：{differences}')

                                        # differences = [S_RemFuelIn_value[i] - S_RemFuelIn_value[i + 1] for i in
                                        #                range(len(S_RemFuelIn_value) - 1)]
                                        # positive_differences = [x for x in differences if x > 0]
                                        # Once_RemFuelIn = round(sum(positive_differences), 2)
                                        # if Once_RemFuelIn == 0:
                                        #     Once_RemFuelIn = 0.3
                                        # Once_S_RemFuelIn.append(Once_RemFuelIn)
                                        # print(f'每次发电消耗燃料（L）:{Once_RemFuelIn}')
                                        # S_RemFuelIn_value.clear()  # 用完S_RemFuelIn_value列表后，要把列表清空，不然会叠加列表

                                        # 计算发电过程中，A电堆电压平均值（过滤小于90和大于130的值）
                                        average_A_StackV = round(calculate_filtered_average(A_StackV_value), 1)
                                        everytime_A_StackV.append(average_A_StackV)
                                        #### 2023.1.16新增
                                        copy_everytime_A_StackV = copy.deepcopy(everytime_A_StackV)
                                        copys_everytime_A_StackV.append(copy_everytime_A_StackV)
                                        modified_A_StackV = [item[0] for item in copys_everytime_A_StackV]
                                        ######
                                        print(f'A电堆平均电压(V):{average_A_StackV}', end="        ")
                                        # print(f'A电堆平均电压  -------- (V):{A_StackV_value}')
                                        A_StackV_value.clear()  # 用完A_StackV_value列表后，要把列表清空，不然会叠加列表
                                        everytime_A_StackV.clear()  # everytime_A_StackV 用于计算平均值。每次算完后列表清零

                                        # 计算发电过程中，B电堆电压平均值（过滤小于90和大于130的值）
                                        # everytime_B_StackV 用于计算平均值。每次算完后列表清零
                                        average_B_StackV = round(calculate_filtered_average(B_StackV_value), 1)
                                        everytime_B_StackV.append(average_B_StackV)
                                        #### 2023.1.16新增
                                        copy_everytime_B_StackV = copy.deepcopy(everytime_B_StackV)
                                        copys_everytime_B_StackV.append(copy_everytime_B_StackV)
                                        modified_B_StackV = [item[0] for item in copys_everytime_B_StackV]
                                        ######
                                        print(f'B电堆平均电压(V):{average_B_StackV}')

                                        # print(f'B电堆平均电压  -------- (V):{B_StackV_value}')

                                        B_StackV_value.clear()  # 用完B_StackV_value列表后，要把列表清空，不然会叠加列表
                                        everytime_B_StackV.clear()  # everytime_B_StackV 用于计算平均值。每次算完后列表清零

                                        if all(item == 0 for item in HGretem_value) and all(
                                                item == 0 for item in Hfetem_value):

                                            max_HGretem = 0
                                            everytime_max_HGretem.append(max_HGretem)
                                            print(f'重整室最大温度(℃)：{max_HGretem}', end="      ")

                                            min_HGretem = 0
                                            everytime_min_HGretem.append(min_HGretem)
                                            print(f'重整室最小温度(℃)：{min_HGretem}')

                                            # print(f'重整室最列表温度^^^^^^^^^^^^5(℃)：{HGretem_value}')
                                            HGretem_value = []  # 用完HGretem_value列表后，要把列表清空，不然会叠加列表

                                            max_Hfetem = 0
                                            everytime_max_Hfetem.append(max_Hfetem)
                                            print(f'提纯器最大温度(℃)：{max_Hfetem}', end="      ")

                                            min_Hfetem = 0
                                            everytime_min_Hfetem.append(min_Hfetem)
                                            print(f'提纯器最小温度(℃)：{min_Hfetem}')

                                            # print(f'提纯器温度列表^^^^^^^^^^^^^^^6(℃)：{Hfetem_value}')
                                            Hfetem_value = []

                                        else:
                                            #   使用列表推导式过滤了列表 HGretem_value 中值为 0 的元素，并将结果重新赋值给 HGretem_value
                                            HGretem_value = [x for x in HGretem_value if x != 0]
                                            max_HGretem = round(max(HGretem_value), 1)
                                            everytime_max_HGretem.append(max_HGretem)
                                            print(f'重整室最大温度(℃)：{max_HGretem}', end="      ")

                                            min_HGretem = round(min(HGretem_value), 1)
                                            everytime_min_HGretem.append(min_HGretem)
                                            print(f'重整室最小温度(℃)：{min_HGretem}')

                                            # print(f'重整室最列表温度^^^^^^^^^^^^5(℃)：{HGretem_value}')
                                            HGretem_value = []  # 用完HGretem_value列表后，要把列表清空，不然会叠加列表

                                            #   使用列表推导式过滤了列表 Hfetem_value 中值为 0 的元素，并将结果重新赋值给 Hfetem_value
                                            Hfetem_value = [x for x in Hfetem_value if x != 0]
                                            max_Hfetem = round(max(Hfetem_value), 1)
                                            everytime_max_Hfetem.append(max_Hfetem)
                                            print(f'提纯器最大温度(℃)：{max_Hfetem}', end="      ")

                                            min_Hfetem = round(min(Hfetem_value), 1)
                                            everytime_min_Hfetem.append(min_Hfetem)
                                            print(f'提纯器最小温度(℃)：{min_Hfetem}')

                                            # print(f'提纯器温度列表^^^^^^^^^^^^^^^6(℃)：{Hfetem_value}')
                                            Hfetem_value = []  # 用完Hfetem_value列表后，要把列表清空，不然会叠加列表

                                        # 燃料耗率 / L.kWh - 1
                                        if Once_Topgen != 0:
                                            Fuel_consumption = round((Once_RemFuelIn / Once_Topgen), 1)
                                        else:
                                            Fuel_consumption = 0
                                        everytime_Fuel_consumption.append(Fuel_consumption)
                                        print(f'燃料消耗率列表 ：{Fuel_consumption}')

                                    # 一天发一次电以上，执行下面程序
                                    if len(count_end_datatime) > 1:

                                        # 找出每次发电期间（内置水箱剩余燃料）fuel_List_value 的所有值
                                        # 求出两个列表长度不同的部分。这段代码使用了 Python 中的切片操作。我们知道，对一个列表进行切片操作时，
                                        # 可以指定起始位置和结束位置，如果只有一个位置（索引），则表示从那个位置到列表末尾。 在这里，fuel_levels[len(last_fuel_levels):] 表示从
                                        # fuel_levels 列表中的索引 len(last_fuel_levels) 开始， 一直取到末尾，即取出 fuel_levels
                                        # 计算液位。 last_fuel_levels 多出来的部分元素。
                                        fuel_List_value = S_RemFuelIn_value[len(last_fuel_levels):]

                                        # 计算电压重整室温度。每次发电期间 HGretem_List_value 重整室温度的值
                                        HGretem_List_value = HGretem_value[len(last_HGretem_list):]
                                        # 计算提纯器温度。每次发电期间 Hfetem_List_value 提纯器温度的值
                                        Hfetem_List_value = Hfetem_value[len(last_Hfetem_list):]
                                        # print(f'上一个power_values功率值>>>>>>>>>>>>：{power_values}')
                                        # print(f'上一个last_power_value_list功率值>>>>>>>>>>>>：{last_power_value_list}')
                                        # 找出每次发电期间（A 电堆电压）A_List_value 的所有值
                                        A_List_value = A_StackV_value[len(last_A_List):]
                                        # 找出每次发电期间（B 电堆电压）A_List_value 的所有值
                                        B_List_value = B_StackV_value[len(last_B_List):]
                                        # 找出每次发电期间（发电功率）last_power_value_list 的所有值
                                        power_value_list = power_values[len(last_power_value_list):]
                                        power_A_value_list = A_Power_values[len(last_A_power_value_list):]
                                        power_B_value_list = B_Power_values[len(last_B_power_value_list):]
                                        # 找出每次发电期间（母线电压）current_voltage_List_value 的所有值
                                        current_voltage_List_value = current_voltage_value[
                                                                     len(last_current_voltage_List_value):]

                                        current_voltage = round(
                                            sum(current_voltage_List_value) / len(current_voltage_List_value), 1)
                                        everytime_current_voltage.append(current_voltage)
                                        current_voltage_List_value.clear()
                                        print(f'母线电压平均值(W)：{current_voltage}')

                                        # print(f'总功率平均值>>>>>>>>>>>>：{power_value_list}')
                                        calculate_A_power = round(calculate_average(power_A_value_list), 1)
                                        everytime_A_power.append(calculate_A_power)
                                        power_A_value_list.clear()
                                        print(f'A堆功率平均值(W)：{calculate_A_power}')

                                        calculate_B_power = round(calculate_average(power_B_value_list), 1)
                                        everytime_B_power.append(calculate_B_power)
                                        power_B_value_list.clear()
                                        print(f'B堆功率平均值(W)：{calculate_B_power}')

                                        # print(f'总功率平均值>>>>>>>>>>>>：{power_value_list}')
                                        calculate_power = round(calculate_average(power_value_list), 1)
                                        everytime_power.append(calculate_power)
                                        power_value_list.clear()
                                        print(f'总功率平均值(W)：{calculate_power}')

                                        if S_RemFuelIn_value[0] > 0:
                                            # 计算燃料使用，计算列表中两两元素的差,大于等于0的部分存到一个新的列表中
                                            differences = [fuel_List_value[i] - fuel_List_value[i + 1] for i in
                                                           range(len(fuel_List_value) - 1)]
                                            positive_differences = [x for x in differences if x > 0]
                                            Once_RemFuelIn = round(sum(positive_differences), 2)
                                            if Once_RemFuelIn == 0:
                                                Once_RemFuelIn = 0.3
                                            Once_S_RemFuelIn.append(Once_RemFuelIn)
                                            print(f'每次发电消耗燃料（L）:{Once_RemFuelIn}')
                                        else:
                                            differences = round(start_LiqlelM[-1] - end_LiqlelM[-1], 1)
                                            if differences < 0:
                                                differences = 0
                                            Once_S_RemFuelIn.append(differences)
                                            print(f'每次发电消耗燃料（mm）:{differences}')
                                            # print(f'液位(mm)******** ：{differences}')

                                        # # 计算燃料使用，计算列表中两两元素的差,大于等于0的部分存到一个新的列表中
                                        # differences = [fuel_List_value[i] - fuel_List_value[i + 1] for i in
                                        #                range(len(fuel_List_value) - 1)]
                                        # positive_differences = [x for x in differences if x > 0]
                                        # Once_RemFuelIn = round(sum(positive_differences), 2)
                                        # if Once_RemFuelIn == 0:
                                        #     Once_RemFuelIn = 0.3
                                        # Once_S_RemFuelIn.append(Once_RemFuelIn)
                                        # print(f'每次发电消耗燃料（L）:{Once_RemFuelIn}')

                                        # 计算发电过程中，A电堆电压平均值（过滤小于90和大于130的值）
                                        average_A_StackV = round(calculate_filtered_average(A_List_value), 1)
                                        everytime_A_StackV.append(average_A_StackV)
                                        #### 2023.1.16新增
                                        copy_everytime_A_StackV = copy.deepcopy(everytime_A_StackV)
                                        copys_everytime_A_StackV.append(copy_everytime_A_StackV)
                                        modified_A_StackV = [item[0] for item in copys_everytime_A_StackV]
                                        everytime_A_StackV.clear()
                                        ######
                                        print(f'A电堆平均电压(V):{average_A_StackV}', end="        ")

                                        # 计算发电过程中，B电堆电压平均值（过滤小于90和大于130的值）
                                        average_B_StackV = round(calculate_filtered_average(B_List_value), 1)
                                        everytime_B_StackV.append(average_B_StackV)
                                        #### 2023.1.16新增
                                        copy_everytime_B_StackV = copy.deepcopy(everytime_B_StackV)
                                        copys_everytime_B_StackV.append(copy_everytime_B_StackV)
                                        modified_B_StackV = [item[0] for item in copys_everytime_B_StackV]
                                        everytime_B_StackV.clear()
                                        ######
                                        print(f'B电堆平均电压(V):{average_B_StackV}')

                                        # print(f'重整室温度 HGretem_List_value ///////////// (℃) ：{HGretem_List_value}')
                                        if all(item == 0 for item in HGretem_List_value) and all(
                                                item == 0 for item in Hfetem_List_value):

                                            max_HGretem = 0
                                            everytime_max_HGretem.append(max_HGretem)
                                            print(f'重整室最大温度(℃)：{max_HGretem}', end="      ")

                                            min_HGretem = 0
                                            everytime_min_HGretem.append(min_HGretem)
                                            print(f'重整室最小温度(℃)：{min_HGretem}')

                                            # print(f'重整室最列表温度^^^^^^^^^^^^5(℃)：{HGretem_value}')
                                            HGretem_value = []  # 用完HGretem_value列表后，要把列表清空，不然会叠加列表

                                            max_Hfetem = 0
                                            everytime_max_Hfetem.append(max_Hfetem)
                                            print(f'提纯器最大温度(℃)：{max_Hfetem}', end="      ")

                                            min_Hfetem = 0
                                            everytime_min_Hfetem.append(min_Hfetem)
                                            print(f'提纯器最小温度(℃)：{min_Hfetem}')

                                            # print(f'提纯器温度列表^^^^^^^^^^^^^^^6(℃)：{Hfetem_value}')
                                            Hfetem_value = []

                                        else:
                                            # print(f'重整室温度列表(℃)>>>>>>>>>>>>>>>>>：{HGretem_List_value}\n')
                                            # print(f'提纯器温度列表(℃)>>>>>>>>>>>>>>>>>：{Hfetem_List_value}\n')

                                            #   使用列表推导式过滤了列表 HGretem_value 中值为 0 的元素，并将结果重新赋值给 HGretem_value
                                            HGretem_List_value = [x for x in HGretem_List_value if x != 0]
                                            max_HGretem = round(max(HGretem_List_value), 1)
                                            everytime_max_HGretem.append(max_HGretem)
                                            print(f'重整室最大温度(℃)：{max_HGretem}', end="      ")

                                            min_HGretem = round(min(HGretem_List_value), 1)
                                            everytime_min_HGretem.append(min_HGretem)
                                            print(f'重整室最小温度(℃)：{min_HGretem}')
                                            # print(f'重整室最温度列表 00000000  (℃)：{HGretem_List_value}')
                                            # print(f'重整室最小温度 HGretem_List_value |||||||||||  (℃)：{HGretem_List_value}')

                                            #   使用列表推导式过滤了列表 Hfetem_value 中值为 0 的元素，并将结果重新赋值给 Hfetem_value
                                            Hfetem_List_value = [x for x in Hfetem_List_value if x != 0]
                                            max_Hfetem = round(max(Hfetem_List_value), 1)
                                            everytime_max_Hfetem.append(max_Hfetem)
                                            print(f'提纯器最大温度(℃)：{max_Hfetem}', end="      ")

                                            min_Hfetem = round(min(Hfetem_List_value), 1)
                                            everytime_min_Hfetem.append(min_Hfetem)
                                            print(f'提纯器最小温度(℃)：{min_Hfetem}')

                                        # 初始化,上一个的列表
                                        last_fuel_levels.clear()
                                        last_A_List.clear()
                                        last_B_List.clear()
                                        last_Hfetem_list.clear()
                                        last_HGretem_list.clear()
                                        last_power_value_list.clear()
                                        last_A_power_value_list.clear()
                                        last_B_power_value_list.clear()
                                        last_current_voltage_List_value.clear()

                                        # 在每次迭代结束后，将 fuel_levels 的值复制给 last_fuel_levels
                                        # 使用 copy 模块中的 deepcopy 函数来创建一个深层副本，确保每个元素都是独立的
                                        # 赋值，将当前列表的值赋于另一个列表，使另一个列表成为上一个列表的值
                                        last_fuel_levels = copy.deepcopy(S_RemFuelIn_value)
                                        last_A_List = copy.deepcopy(A_StackV_value)
                                        last_B_List = copy.deepcopy(B_StackV_value)
                                        last_HGretem_list = copy.deepcopy(HGretem_value)
                                        last_Hfetem_list = copy.deepcopy(Hfetem_value)
                                        last_power_value_list = copy.deepcopy(power_values)
                                        last_A_power_value_list = copy.deepcopy(A_Power_values)
                                        last_B_power_value_list = copy.deepcopy(B_Power_values)
                                        last_current_voltage_List_value = copy.deepcopy(current_voltage_value)

                                        # 燃料耗率 / L.kWh - 1
                                        if Once_Topgen != 0:
                                            Fuel_consumption = round((Once_RemFuelIn / Once_Topgen), 1)
                                        else:
                                            Fuel_consumption = 0
                                        everytime_Fuel_consumption.append(Fuel_consumption)
                                        print(f'燃料消耗率列表 ：{Fuel_consumption}')

                                    print('=============     本次发电结束      ==================')

                                else:
                                    Once_RemFuelIn = 0
                                    if start_time is None and (index == len(df) - 1) == True and last_row[
                                        MSw] == True and len(
                                        count_end_datatime) == 1:  # 有开始发电时间并且到列的最后一行，把最后一行的数值添加进去
                                        print(
                                            f"结束发电时间：{row[DateTime]}      "
                                            f"内置水箱剩余燃料: {round(row[S_RemFuelIn], 2)}     "
                                            f"外置水箱剩余燃料: {round(row[S_RemFuelOut], 2)}    "
                                            f"内置水箱剩余燃料(mm): {round(row[LiqlelM], 1)} "
                                            f"外置水箱剩余燃料(mm): {round(row[LiqlelL], 1)}"
                                            f"总发电量:{row[Topgen]}    ")

                                        print(len(count_end_datatime))  # 计算当天发电次数
                                        Time_value.append(row[DateTime])
                                        end_datatime.append(row[DateTime])
                                        Topgen_value.append(row[Topgen])
                                        # 创建列表用于储存输出到excel表格和数据

                                        # 创建列表count_end_datatime，用于计数。一天发了多少次电

                                        end_S_RemFuelIn.append(round(row[S_RemFuelIn], 1))
                                        end_Topgen.append(round(row[Topgen], 1))
                                        end_S_RemFuelOut.append(round(row[S_RemFuelOut], 1))

                                        Once_Topgen = round(Topgen_value[-1] - Topgen_value[-2], 3)
                                        print(f"每次发电量(kw/h)：{Once_Topgen}")
                                        Once_Topgen_value.append(Once_Topgen)

                                        Stwtims.append(row[Stwtim])
                                        print(f"发电次数：{row[Stwtim]}")

                                        Time_diff = round(
                                            (pd.to_datetime(Time_value[-1]) - pd.to_datetime(
                                                Time_value[-2])).total_seconds() / 60,
                                            2)
                                        Time_diffs.append(Time_diff)
                                        print(f"每次发电时长(min)：{Time_diff}")

                                        mean_IC = round(sum(IC_value) / len(IC_value), 2)
                                        everytime_IC.append(mean_IC)
                                        print(f'芯片平均温度(℃):{mean_IC}')

                                        # 计算液位。 last_fuel_levels 多出来的部分元素。
                                        fuel_List_value = S_RemFuelIn_value[len(last_fuel_levels):]
                                        # print(f'最后一次液位 >>>>>>>>>>>>：{fuel_List_value} \n')
                                        # 计算电压重整室温度。每次发电期间 HGretem_List_value 重整室温度的值
                                        HGretem_List_value = HGretem_value[len(last_HGretem_list):]
                                        # print(f'最后一次重整室温度 >>>>>>>>>>>>：{HGretem_List_value} \n')
                                        # 计算提纯器温度。每次发电期间 Hfetem_List_value 提纯器温度的值
                                        Hfetem_List_value = Hfetem_value[len(last_Hfetem_list):]
                                        # print(f'最后一次提纯器温度 >>>>>>>>>>>>：{Hfetem_List_value} \n')
                                        # 找出每次发电期间（A 电堆电压）A_List_value 的所有值
                                        A_List_value = A_StackV_value[len(last_A_List):]
                                        # print(f'最后一次 A 电堆电压 >>>>>>>>>>>>：{A_List_value} \n')
                                        # 找出每次发电期间（B 电堆电压）A_List_value 的所有值
                                        B_List_value = B_StackV_value[len(last_B_List):]
                                        # print(f'最后一次 B 电堆电压 >>>>>>>>>>>>：{B_List_value} \n')
                                        # 找出每次发电期间（发电功率）last_power_value_list 的所有值
                                        power_value_list = power_values[len(last_power_value_list):]
                                        # print(f'总功率平均值>>>>>>>>>>>>：{power_value_list} \n')

                                        power_A_value_list = A_Power_values[len(last_A_power_value_list):]
                                        power_B_value_list = B_Power_values[len(last_B_power_value_list):]

                                        current_voltage_List_value = current_voltage_value[
                                                                     len(last_current_voltage_List_value):]

                                        current_voltage = round(
                                            sum(current_voltage_List_value) / len(current_voltage_List_value), 1)
                                        everytime_current_voltage.append(current_voltage)
                                        current_voltage_List_value.clear()
                                        print(f'母线电压平均值(W)：{current_voltage}')

                                        # print(f'总功率平均值>>>>>>>>>>>>：{power_value_list}')
                                        calculate_A_power = round(calculate_average(power_A_value_list), 1)
                                        everytime_A_power.append(calculate_A_power)
                                        power_A_value_list.clear()
                                        print(f'A堆功率平均值(W)：{calculate_A_power}')

                                        calculate_B_power = round(calculate_average(power_B_value_list), 1)
                                        everytime_B_power.append(calculate_B_power)
                                        power_B_value_list.clear()
                                        print(f'B堆功率平均值(W)：{calculate_B_power}')

                                        calculate_power = round(calculate_average(power_value_list), 1)
                                        everytime_power.append(calculate_power)
                                        power_value_list.clear()
                                        print(f'总功率平均值(W)：{calculate_power}')

                                        if S_RemFuelIn_value[0] > 0:
                                            # 计算燃料使用，计算列表中两两元素的差,大于等于0的部分存到一个新的列表中
                                            differences = [fuel_List_value[i] - fuel_List_value[i + 1] for i in
                                                           range(len(fuel_List_value) - 1)]
                                            positive_differences = [x for x in differences if x > 0]
                                            Once_RemFuelIn = round(sum(positive_differences), 2)
                                            if Once_RemFuelIn == 0:
                                                Once_RemFuelIn = 0.3
                                            Once_S_RemFuelIn.append(Once_RemFuelIn)
                                            print(f'每次发电消耗燃料（L）:{Once_RemFuelIn}')
                                        else:
                                            differences = round(start_LiqlelM[-1] - end_LiqlelM[-1], 1)
                                            if differences < 0:
                                                differences = 0
                                            Once_S_RemFuelIn.append(differences)
                                            print(f'每次发电消耗燃料（mm）:{differences}')

                                        # # 计算燃料使用，计算列表中两两元素的差,大于等于0的部分存到一个新的列表中
                                        # differences = [fuel_List_value[i] - fuel_List_value[i + 1] for i in
                                        #                range(len(fuel_List_value) - 1)]
                                        # positive_differences = [x for x in differences if x > 0]
                                        # Once_RemFuelIn = round(sum(positive_differences), 2)
                                        # if Once_RemFuelIn == 0:
                                        #     Once_RemFuelIn = 0.3
                                        # Once_S_RemFuelIn.append(Once_RemFuelIn)
                                        # print(f'每次发电消耗燃料（L）:{Once_RemFuelIn}')

                                        # 计算发电过程中，A电堆电压平均值（过滤小于90和大于130的值）
                                        average_A_StackV = round(calculate_filtered_average(A_List_value), 1)
                                        everytime_A_StackV.append(average_A_StackV)
                                        #### 2023.1.16新增
                                        copy_everytime_A_StackV = copy.deepcopy(everytime_A_StackV)
                                        copys_everytime_A_StackV.append(copy_everytime_A_StackV)
                                        modified_A_StackV = [item[0] for item in copys_everytime_A_StackV]
                                        everytime_A_StackV.clear()
                                        ######
                                        print(f'A电堆平均电压(V):{average_A_StackV}', end="        ")

                                        # 计算发电过程中，B电堆电压平均值（过滤小于90和大于130的值）
                                        average_B_StackV = round(calculate_filtered_average(B_List_value), 1)
                                        everytime_B_StackV.append(average_B_StackV)
                                        #### 2023.1.16新增
                                        copy_everytime_B_StackV = copy.deepcopy(everytime_B_StackV)
                                        copys_everytime_B_StackV.append(copy_everytime_B_StackV)
                                        modified_B_StackV = [item[0] for item in copys_everytime_B_StackV]
                                        everytime_B_StackV.clear()
                                        ######
                                        print(f'B电堆平均电压(V):{average_B_StackV}')

                                        # print(f'重整室温度 HGretem_List_value ///////////// (℃) ：{HGretem_List_value}')
                                        if all(item == 0 for item in HGretem_List_value) and all(
                                                item == 0 for item in Hfetem_List_value):

                                            max_HGretem = 0
                                            everytime_max_HGretem.append(max_HGretem)
                                            print(f'重整室最大温度(℃)：{max_HGretem}', end="      ")

                                            min_HGretem = 0
                                            everytime_min_HGretem.append(min_HGretem)
                                            print(f'重整室最小温度(℃)：{min_HGretem}')

                                            # print(f'重整室最列表温度^^^^^^^^^^^^5(℃)：{HGretem_value}')
                                            HGretem_value = []  # 用完HGretem_value列表后，要把列表清空，不然会叠加列表

                                            max_Hfetem = 0
                                            everytime_max_Hfetem.append(max_Hfetem)
                                            print(f'提纯器最大温度(℃)：{max_Hfetem}', end="      ")

                                            min_Hfetem = 0
                                            everytime_min_Hfetem.append(min_Hfetem)
                                            print(f'提纯器最小温度(℃)：{min_Hfetem}')

                                            # print(f'提纯器温度列表^^^^^^^^^^^^^^^6(℃)：{Hfetem_value}')
                                            Hfetem_value = []

                                        else:
                                            # print(f'重整室温度列表(℃)>>>>>>>>>>>>>>>>>：{HGretem_List_value}\n')
                                            # print(f'提纯器温度列表(℃)>>>>>>>>>>>>>>>>>：{Hfetem_List_value}\n')

                                            #   使用列表推导式过滤了列表 HGretem_value 中值为 0 的元素，并将结果重新赋值给 HGretem_value
                                            HGretem_List_value = [x for x in HGretem_List_value if x != 0]
                                            max_HGretem = round(max(HGretem_List_value), 1)
                                            everytime_max_HGretem.append(max_HGretem)
                                            print(f'重整室最大温度(℃)：{max_HGretem}', end="      ")

                                            min_HGretem = round(min(HGretem_List_value), 1)
                                            everytime_min_HGretem.append(min_HGretem)
                                            print(f'重整室最小温度(℃)：{min_HGretem}')
                                            # print(f'重整室最温度列表 00000000  (℃)：{HGretem_List_value}')
                                            # print(f'重整室最小温度 HGretem_List_value |||||||||||  (℃)：{HGretem_List_value}')

                                            #   使用列表推导式过滤了列表 Hfetem_value 中值为 0 的元素，并将结果重新赋值给 Hfetem_value
                                            Hfetem_List_value = [x for x in Hfetem_List_value if x != 0]
                                            max_Hfetem = round(max(Hfetem_List_value), 1)
                                            everytime_max_Hfetem.append(max_Hfetem)
                                            print(f'提纯器最大温度(℃)：{max_Hfetem}', end="      ")

                                            min_Hfetem = round(min(Hfetem_List_value), 1)
                                            everytime_min_Hfetem.append(min_Hfetem)
                                            print(f'提纯器最小温度(℃)：{min_Hfetem}')

                                        # 燃料耗率 / L.kWh - 1
                                        if Once_Topgen != 0:
                                            Fuel_consumption = round((Once_RemFuelIn / Once_Topgen), 1)
                                        else:
                                            Fuel_consumption = 0
                                        everytime_Fuel_consumption.append(Fuel_consumption)
                                        print(f'燃料消耗率 ：{Fuel_consumption}')

                                        # 初始化,上一个的列表
                                        last_fuel_levels.clear()
                                        last_A_List.clear()
                                        last_B_List.clear()
                                        last_Hfetem_list.clear()
                                        last_HGretem_list.clear()
                                        last_power_value_list.clear()
                                        last_A_power_value_list.clear()
                                        last_B_power_value_list.clear()
                                        last_current_voltage_List_value.clear()

                                        # 在每次迭代结束后，将 fuel_levels 的值复制给 last_fuel_levels
                                        # 使用 copy 模块中的 deepcopy 函数来创建一个深层副本，确保每个元素都是独立的
                                        # 赋值，将当前列表的值赋于另一个列表，使另一个列表成为上一个列表的值
                                        last_fuel_levels = copy.deepcopy(S_RemFuelIn_value)
                                        last_A_List = copy.deepcopy(A_StackV_value)
                                        last_B_List = copy.deepcopy(B_StackV_value)
                                        last_HGretem_list = copy.deepcopy(HGretem_value)
                                        last_Hfetem_list = copy.deepcopy(Hfetem_value)
                                        last_power_value_list = copy.deepcopy(power_values)
                                        last_A_power_value_list = copy.deepcopy(A_Power_values)
                                        last_B_power_value_list = copy.deepcopy(B_Power_values)
                                        last_current_voltage_List_value = copy.deepcopy(current_voltage_value)

                                    if start_time is None and (index == len(df) - 1) == True and last_row[
                                        MSw] == True and len(
                                        count_end_datatime) > 1:
                                        print(
                                            f"结束发电时间：{last_row[DateTime]}      "
                                            f"内置水箱剩余燃料: {round(last_row[S_RemFuelIn], 2)}     "
                                            f"外置水箱剩余燃料: {round(last_row[S_RemFuelOut], 2)}    "
                                            f"内置水箱剩余燃料(mm): {round(prev_row[LiqlelM], 1)} "
                                            f"外置水箱剩余燃料(mm): {round(prev_row[LiqlelL], 1)}"
                                            f"总发电量:{last_row[Topgen]}    ")

                                        print(len(count_end_datatime))  # 计算当天发电次数
                                        Time_value.append(last_row[DateTime])
                                        end_datatime.append(last_row[DateTime])
                                        Topgen_value.append(last_row[Topgen])
                                        # 创建列表用于储存输出到excel表格和数据

                                        # 创建列表count_end_datatime，用于计数。一天发了多少次电

                                        end_S_RemFuelIn.append(round(last_row[S_RemFuelIn], 1))
                                        end_Topgen.append(round(last_row[Topgen], 1))
                                        end_S_RemFuelOut.append(round(last_row[S_RemFuelOut], 1))

                                        Once_Topgen = round(Topgen_value[-1] - Topgen_value[-2], 3)
                                        print(f"每次发电量(kw/h)：{Once_Topgen}")
                                        Once_Topgen_value.append(Once_Topgen)

                                        Stwtims.append(last_row[Stwtim])
                                        print(f"发电次数：{row[Stwtim]}")

                                        Time_diff = round(
                                            (pd.to_datetime(Time_value[-1]) - pd.to_datetime(
                                                Time_value[-2])).total_seconds() / 60,
                                            2)
                                        Time_diffs.append(Time_diff)
                                        print(f"每次发电时长(min)：{Time_diff}")

                                        mean_IC = round(sum(IC_value) / len(IC_value), 2)
                                        everytime_IC.append(mean_IC)
                                        print(f'芯片平均温度(℃):{mean_IC}')

                                        # 计算液位。 last_fuel_levels 多出来的部分元素。
                                        fuel_List_value = S_RemFuelIn_value[len(last_fuel_levels):]
                                        # print(f'最后一次液位 >>>>>>>>>>>>：{fuel_List_value} \n')
                                        # 计算电压重整室温度。每次发电期间 HGretem_List_value 重整室温度的值
                                        HGretem_List_value = HGretem_value[len(last_HGretem_list):]
                                        # print(f'最后一次重整室温度 >>>>>>>>>>>>：{HGretem_List_value} \n')
                                        # 计算提纯器温度。每次发电期间 Hfetem_List_value 提纯器温度的值
                                        Hfetem_List_value = Hfetem_value[len(last_Hfetem_list):]
                                        # print(f'最后一次提纯器温度 >>>>>>>>>>>>：{Hfetem_List_value} \n')
                                        # 找出每次发电期间（A 电堆电压）A_List_value 的所有值
                                        A_List_value = A_StackV_value[len(last_A_List):]
                                        # print(f'最后一次 A 电堆电压 >>>>>>>>>>>>：{A_List_value} \n')
                                        # 找出每次发电期间（B 电堆电压）A_List_value 的所有值
                                        B_List_value = B_StackV_value[len(last_B_List):]
                                        # print(f'最后一次 B 电堆电压 >>>>>>>>>>>>：{B_List_value} \n')
                                        # 找出每次发电期间（发电功率）last_power_value_list 的所有值
                                        power_value_list = power_values[len(last_power_value_list):]
                                        # print(f'总功率平均值>>>>>>>>>>>>：{power_value_list} \n')
                                        current_voltage_List_value = current_voltage_value[
                                                                     len(last_current_voltage_List_value):]

                                        current_voltage = round(
                                            sum(current_voltage_List_value) / len(current_voltage_List_value), 1)
                                        everytime_current_voltage.append(current_voltage)
                                        current_voltage_List_value.clear()
                                        print(f'母线电压平均值(W)：{current_voltage}')

                                        power_A_value_list = A_Power_values[len(last_A_power_value_list):]
                                        power_B_value_list = B_Power_values[len(last_B_power_value_list):]

                                        # print(f'总功率平均值>>>>>>>>>>>>：{power_value_list}')
                                        calculate_A_power = round(calculate_average(power_A_value_list), 1)
                                        everytime_A_power.append(calculate_A_power)
                                        power_A_value_list.clear()
                                        print(f'A堆功率平均值(W)：{calculate_A_power}')

                                        calculate_B_power = round(calculate_average(power_B_value_list), 1)
                                        everytime_B_power.append(calculate_B_power)
                                        power_B_value_list.clear()
                                        print(f'B堆功率平均值(W)：{calculate_B_power}')

                                        calculate_power = round(calculate_average(power_value_list), 1)
                                        everytime_power.append(calculate_power)
                                        power_value_list.clear()
                                        print(f'总功率平均值(W)：{calculate_power}')

                                        if S_RemFuelIn_value[0] > 0:
                                            # 计算燃料使用，计算列表中两两元素的差,大于等于0的部分存到一个新的列表中
                                            differences = [fuel_List_value[i] - fuel_List_value[i + 1] for i in
                                                           range(len(fuel_List_value) - 1)]
                                            positive_differences = [x for x in differences if x > 0]
                                            Once_RemFuelIn = round(sum(positive_differences), 2)
                                            if Once_RemFuelIn == 0:
                                                Once_RemFuelIn = 0.3
                                            Once_S_RemFuelIn.append(Once_RemFuelIn)
                                            print(f'每次发电消耗燃料（L）:{Once_RemFuelIn}')
                                        else:
                                            differences = round(start_LiqlelM[-1] - end_LiqlelM[-1], 1)
                                            if differences < 0:
                                                differences = 0
                                            Once_S_RemFuelIn.append(differences)
                                            print(f'每次发电消耗燃料（mm）:{differences}')
                                            # print(f'液位(mm)******** ：{differences}')

                                        # # 计算燃料使用，计算列表中两两元素的差,大于等于0的部分存到一个新的列表中
                                        # differences = [fuel_List_value[i] - fuel_List_value[i + 1] for i in
                                        #                range(len(fuel_List_value) - 1)]
                                        # positive_differences = [x for x in differences if x > 0]
                                        # Once_RemFuelIn = round(sum(positive_differences), 2)
                                        # if Once_RemFuelIn == 0:
                                        #     Once_RemFuelIn = 0.3
                                        # Once_S_RemFuelIn.append(Once_RemFuelIn)
                                        # print(f'每次发电消耗燃料（L）:{Once_RemFuelIn}')

                                        # 计算发电过程中，A电堆电压平均值（过滤小于90和大于130的值）
                                        average_A_StackV = round(calculate_filtered_average(A_List_value), 1)
                                        everytime_A_StackV.append(average_A_StackV)
                                        #### 2023.1.16新增
                                        copy_everytime_A_StackV = copy.deepcopy(everytime_A_StackV)
                                        copys_everytime_A_StackV.append(copy_everytime_A_StackV)
                                        modified_A_StackV = [item[0] for item in copys_everytime_A_StackV]
                                        everytime_A_StackV.clear()
                                        ######
                                        print(f'A电堆平均电压(V):{average_A_StackV}', end="        ")

                                        # 计算发电过程中，B电堆电压平均值（过滤小于90和大于130的值）
                                        average_B_StackV = round(calculate_filtered_average(B_List_value), 1)
                                        everytime_B_StackV.append(average_B_StackV)
                                        #### 2023.1.16新增
                                        copy_everytime_B_StackV = copy.deepcopy(everytime_B_StackV)
                                        copys_everytime_B_StackV.append(copy_everytime_B_StackV)
                                        modified_B_StackV = [item[0] for item in copys_everytime_B_StackV]
                                        everytime_B_StackV.clear()
                                        ######
                                        print(f'B电堆平均电压(V):{average_B_StackV}')

                                        # print(f'重整室温度 HGretem_List_value ///////////// (℃) ：{HGretem_List_value}')
                                        if all(item == 0 for item in HGretem_List_value) and all(
                                                item == 0 for item in Hfetem_List_value):

                                            max_HGretem = 0
                                            everytime_max_HGretem.append(max_HGretem)
                                            print(f'重整室最大温度(℃)：{max_HGretem}', end="      ")

                                            min_HGretem = 0
                                            everytime_min_HGretem.append(min_HGretem)
                                            print(f'重整室最小温度(℃)：{min_HGretem}')

                                            # print(f'重整室最列表温度^^^^^^^^^^^^5(℃)：{HGretem_value}')
                                            HGretem_value = []  # 用完HGretem_value列表后，要把列表清空，不然会叠加列表

                                            max_Hfetem = 0
                                            everytime_max_Hfetem.append(max_Hfetem)
                                            print(f'提纯器最大温度(℃)：{max_Hfetem}', end="      ")

                                            min_Hfetem = 0
                                            everytime_min_Hfetem.append(min_Hfetem)
                                            print(f'提纯器最小温度(℃)：{min_Hfetem}')

                                            # print(f'提纯器温度列表^^^^^^^^^^^^^^^6(℃)：{Hfetem_value}')
                                            Hfetem_value = []

                                        else:
                                            # print(f'重整室温度列表(℃)>>>>>>>>>>>>>>>>>：{HGretem_List_value}\n')
                                            # print(f'提纯器温度列表(℃)>>>>>>>>>>>>>>>>>：{Hfetem_List_value}\n')

                                            #   使用列表推导式过滤了列表 HGretem_value 中值为 0 的元素，并将结果重新赋值给 HGretem_value
                                            HGretem_List_value = [x for x in HGretem_List_value if x != 0]
                                            max_HGretem = round(max(HGretem_List_value), 1)
                                            everytime_max_HGretem.append(max_HGretem)
                                            print(f'重整室最大温度(℃)：{max_HGretem}', end="      ")

                                            min_HGretem = round(min(HGretem_List_value), 1)
                                            everytime_min_HGretem.append(min_HGretem)
                                            print(f'重整室最小温度(℃)：{min_HGretem}')
                                            # print(f'重整室最温度列表 00000000  (℃)：{HGretem_List_value}')
                                            # print(f'重整室最小温度 HGretem_List_value |||||||||||  (℃)：{HGretem_List_value}')

                                            #   使用列表推导式过滤了列表 Hfetem_value 中值为 0 的元素，并将结果重新赋值给 Hfetem_value
                                            Hfetem_List_value = [x for x in Hfetem_List_value if x != 0]
                                            max_Hfetem = round(max(Hfetem_List_value), 1)
                                            everytime_max_Hfetem.append(max_Hfetem)
                                            print(f'提纯器最大温度(℃)：{max_Hfetem}', end="      ")

                                            min_Hfetem = round(min(Hfetem_List_value), 1)
                                            everytime_min_Hfetem.append(min_Hfetem)
                                            print(f'提纯器最小温度(℃)：{min_Hfetem}')

                                        # 燃料耗率 / L.kWh - 1
                                        if Once_Topgen != 0:
                                            Fuel_consumption = round((Once_RemFuelIn / Once_Topgen), 1)
                                        else:
                                            Fuel_consumption = 0
                                        everytime_Fuel_consumption.append(Fuel_consumption)
                                        print(f'燃料消耗率 ：{Fuel_consumption}')

                                        # 初始化,上一个的列表
                                        last_fuel_levels.clear()
                                        last_A_List.clear()
                                        last_B_List.clear()
                                        last_Hfetem_list.clear()
                                        last_HGretem_list.clear()
                                        last_power_value_list.clear()
                                        last_A_power_value_list.clear()
                                        last_B_power_value_list.clear()
                                        last_current_voltage_List_value.clear()

                                        # 在每次迭代结束后，将 fuel_levels 的值复制给 last_fuel_levels
                                        # 使用 copy 模块中的 deepcopy 函数来创建一个深层副本，确保每个元素都是独立的
                                        # 赋值，将当前列表的值赋于另一个列表，使另一个列表成为上一个列表的值
                                        last_fuel_levels = copy.deepcopy(S_RemFuelIn_value)
                                        last_A_List = copy.deepcopy(A_StackV_value)
                                        last_B_List = copy.deepcopy(B_StackV_value)
                                        last_HGretem_list = copy.deepcopy(HGretem_value)
                                        last_Hfetem_list = copy.deepcopy(Hfetem_value)
                                        last_power_value_list = copy.deepcopy(power_values)
                                        last_A_power_value_list = copy.deepcopy(A_Power_values)
                                        last_B_power_value_list = copy.deepcopy(B_Power_values)
                                        last_current_voltage_List_value = copy.deepcopy(current_voltage_value)

                                    start_time = None
                            prev_row = row

                        Sum_Topgen = round(sum(Once_Topgen_value), 2)
                        Sum_S_RemFuelIn = sum(Once_S_RemFuelIn)
                        Sum_Time_min = round(sum(Time_diffs), 2)

                        print(f"总发电量(kw/h)：{Sum_Topgen}")
                        print(f"总发电时间(min.s)：{Sum_Time_min}")

                        if start_S_RemFuelIn[0] > 0:
                            print(f"总燃料消耗(L)：{Sum_S_RemFuelIn}")
                        else:
                            print(f"总燃料消耗(mm)：{Sum_S_RemFuelIn}")

                        # 计数清零，用于计算有多少个【'结束发电时间': end_datatime】。来判断一天里面发了多少次电
                        count_end_datatime.clear()
                        S_RemFuelIn_value.clear()
                        A_StackV_value.clear()
                        B_StackV_value.clear()
                        current_voltage_value.clear()
                        A_Power_values = []
                        B_Power_values = []
                        power_values = []
                        HGretem_value = []
                        Hfetem_value = []
                        last_HGretem_list = []  # 确保在每次循环开始时重置为空列表
                        last_Hfetem_list = []  # 确保在每次循环开始时重置为空列表
                        last_B_List = []
                        last_A_List = []
                        last_fuel_levels = []
                        last_A_power_value_list = []
                        last_B_power_value_list = []
                        last_power_value_list = []
                        count_datatime = []
                        start_time = None
                        second_start_time = None
                        second_row = None
                        first_start_datatime = 0
                        second_end_datatime = 0

                        print(f"\n开始发电时间 长度：{len(start_datatime)}")
                        print(f"结束发电时间 长度：{len(end_datatime)}")
                        print(f"开始外置水箱剩余燃料 长度：{len(start_S_RemFuelOut)}")
                        print(f"结束外置水箱剩余燃料 长度：{len(end_S_RemFuelOut)}")
                        print(f"开始内置水箱剩余燃料 长度：{len(start_S_RemFuelIn)}")
                        print(f"结束内置水箱剩余燃料 长度：{len(end_S_RemFuelIn)}")
                        print(f"开始总发电量 长度：{len(start_Topgen)}")
                        print(f"结束总发电量 长度：{len(end_Topgen)}")
                        print(f"发电功率 长度：{len(everytime_power)}")
                        print(f"芯片温度 长度：{len(everytime_IC)}")
                        print(f"A电堆电压 长度：{len(modified_A_StackV)}")
                        print(f"B电堆电压 长度：{len(modified_B_StackV)}")
                        print(f"重整室最高温度 长度：{len(everytime_max_HGretem)}")
                        print(f"重整室最低温度 长度：{len(everytime_min_HGretem)}")
                        print(f"提纯器最高温度 长度：{len(everytime_max_Hfetem)}")
                        print(f"提纯器最低温度 长度：{len(everytime_min_Hfetem)}")
                        print(f"发电运行时间 长度：{len(Time_diffs)}")
                        print(f"消耗燃料 长度：{len(Once_S_RemFuelIn)}")
                        print(f"发电量 长度：{len(Once_Topgen_value)}")
                        print(f"发电次数 长度：{len(Stwtims)}")
                        print(f"燃料消耗率 长度：{len(everytime_Fuel_consumption)}\n")
                        print(f"母线电压 长度：{len(everytime_current_voltage)}\n")

                        print(f"开始外置水箱剩余燃料(mm) 长度：{len(start_LiqlelL)}")
                        print(f"结束外置水箱剩余燃料(mm) 长度：{len(end_LiqlelL)}")
                        print(f"开始内置水箱剩余燃料(mm) 长度：{len(start_LiqlelM)}")
                        print(f"结束内置水箱剩余燃料(mm) 长度：{len(end_LiqlelM)}")

                        print(f'\n++++++++++++++   一天的计算结束   ++++++++++++++++++++++++\n')

                    else:
                        system_state.append(adress1)
                        # 完成所有操作后更新进度条到100%

                        # 表格页合并完成后，重新启用“发电数据处理”按钮 。条件为假时，执行下面代码
                        if not self.excel_process_button_disabled:
                            self.excel_process_button.config(state=tk.NORMAL)

                        # 表格页合并完成后，重新启用“发电数据处理”按钮 。条件为假时，执行下面代码
                        if not self.no_process_button_disabled:
                            self.no_process_button.config(state=tk.NORMAL)

                        for _ in range(10):
                            self.progress['value'] += 1  # 逐步增加进度条值
                            self.progress.update()
                            time.sleep(0.05)  # 微小的延迟，实现平滑更新
                        # 停止进度条动画
                        self.progress.stop()

                        print(f'\n++++++++++++++   {adress1}    当天没有发电     ++++++++++++++++++++++++\n')

                    # 完成所有操作后更新进度条到100%
                    for _ in range(1):
                        self.progress['value'] += 1  # 逐步增加进度条值
                        self.progress.update()
                        time.sleep(0.05)  # 微小的延迟，实现平滑更新

                except FileNotFoundError:
                    print(f"文件 {adress1} 不存在，已跳过")
            else:
                print(f"文件 {adress1} 不存在，已跳过")
            if len(system_state) == 1:
                self.show_save_fail_message(adress1)
            else:

                if start_S_RemFuelIn[0] > 0 and end_S_RemFuelIn[0] > 0:
                    # 将新的DataFrame保存到新的Excel文件中
                    new_df = pd.DataFrame(
                        {
                            '开始发电时间': start_datatime,
                            '结束发电时间': end_datatime,

                            '开始外置水箱剩余燃料(mm)': start_LiqlelL,
                            '结束外置水箱剩余燃料(mm)': end_LiqlelL,
                            '开始内置水箱剩余燃料(mm)': start_LiqlelM,
                            '结束内置水箱剩余燃料(mm)': end_LiqlelM,

                            '开始外置水箱剩余燃料(L)': start_S_RemFuelOut,
                            '结束外置水箱剩余燃料(L)': end_S_RemFuelOut,
                            '开始内置水箱剩余燃料(L)': start_S_RemFuelIn,
                            '结束内置水箱剩余燃料(L)': end_S_RemFuelIn,
                            '开始总发电量(kw/h)': start_Topgen,
                            '结束总发电量(kw/h)': end_Topgen,
                            '母线电压(V)': everytime_current_voltage,
                            '总发电功率(W)': everytime_power,
                            'A电堆功率(W)': everytime_A_power,
                            'B电堆功率(W)': everytime_B_power,
                            '芯片温度(℃)': everytime_IC,
                            'A电堆电压(V)': modified_A_StackV,
                            'B电堆电压(V)': modified_B_StackV,
                            '重整室最高温度(℃)': everytime_max_HGretem,
                            '重整室最低温度(℃)': everytime_min_HGretem,
                            '提纯器最高温度(℃)': everytime_max_Hfetem,
                            '提纯器最低温度(℃)': everytime_min_Hfetem,
                            '发电运行时间(min.s)': Time_diffs,
                            '消耗燃料(L)': Once_S_RemFuelIn,
                            '发电量(kw/h)': Once_Topgen_value,
                            '发电次数': Stwtims,
                            '燃料消耗率(L.kWh -1)': everytime_Fuel_consumption

                        })

                else:
                    # 将新的DataFrame保存到新的Excel文件中
                    new_df = pd.DataFrame(
                        {
                            '开始发电时间': start_datatime,
                            '结束发电时间': end_datatime,

                            '开始外置水箱剩余燃料(mm)': start_LiqlelL,
                            '结束外置水箱剩余燃料(mm)': end_LiqlelL,
                            '开始内置水箱剩余燃料(mm)': start_LiqlelM,
                            '结束内置水箱剩余燃料(mm)': end_LiqlelM,

                            '开始外置水箱剩余燃料(L)': start_S_RemFuelOut,
                            '结束外置水箱剩余燃料(L)': end_S_RemFuelOut,
                            '开始内置水箱剩余燃料(L)': start_S_RemFuelIn,
                            '结束内置水箱剩余燃料(L)': end_S_RemFuelIn,
                            '开始总发电量(kw/h)': start_Topgen,
                            '结束总发电量(kw/h)': end_Topgen,
                            '母线电压(V)': everytime_current_voltage,
                            '总发电功率(W)': everytime_power,
                            'A电堆功率(W)': everytime_A_power,
                            'B电堆功率(W)': everytime_B_power,
                            '芯片温度(℃)': everytime_IC,
                            'A电堆电压(V)': modified_A_StackV,
                            'B电堆电压(V)': modified_B_StackV,
                            '重整室最高温度(℃)': everytime_max_HGretem,
                            '重整室最低温度(℃)': everytime_min_HGretem,
                            '提纯器最高温度(℃)': everytime_max_Hfetem,
                            '提纯器最低温度(℃)': everytime_min_Hfetem,
                            '发电运行时间(min.s)': Time_diffs,
                            '消耗燃料(mm)': Once_S_RemFuelIn,
                            '发电量(kw/h)': Once_Topgen_value,
                            '发电次数': Stwtims,
                            '燃料消耗率(L.kWh -1)': everytime_Fuel_consumption

                        })

                excel_file_path = adress3

                new_df.to_excel(excel_file_path, index=False, engine='openpyxl')
                # 打开现有的Excel文件
                workbook = openpyxl.load_workbook(excel_file_path)
                # 选择第一个工作表
                sheet = workbook.active
                # 设置第一行的行高
                sheet.row_dimensions[1].height = 50
                # 设置第一列和第二列的宽度为 25
                sheet.column_dimensions['A'].width = 21  # 第一列
                sheet.column_dimensions['B'].width = 21  # 第二列
                # 设置其余列的宽度为 10
                for col in sheet.columns:
                    if col[0].column_letter not in ['A', 'B']:
                        sheet.column_dimensions[col[0].column_letter].width = 10
                # 遍历第一行的所有单元格，并为每个单元格对象同时设置自动换行、水平居中和垂直居中。
                for cell in sheet[1]:
                    cell_obj = cell
                    cell_obj.alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal='center',
                                                                   vertical='center')

                workbook.save(excel_file_path)

                # 完成所有操作后更新进度条到100%
                for _ in range(5):
                    self.progress['value'] += 1  # 逐步增加进度条值
                    self.progress.update()
                    time.sleep(0.05)  # 微小的延迟，实现平滑更新

                # 表格页合并完成后，重新启用“发电数据处理”按钮 。条件为假时，执行下面代码
                if not self.excel_process_button_disabled:
                    self.excel_process_button.config(state=tk.NORMAL)

                # 表格页合并完成后，重新启用“发电数据处理”按钮 。条件为假时，执行下面代码
                if not self.no_process_button_disabled:
                    self.no_process_button.config(state=tk.NORMAL)

                # 停止进度条动画
                self.progress.stop()

                print(f"\n文件保存成功 ！! ! ")
                print(f"文件保存路径 ：{excel_file_path}")
                self.show_save_success_message(adress3)

            system_state.clear()
        except ValueError:

            self.show_read_error(adress1)

            for _ in range(105):
                self.progress['value'] += 1  # 如果，异常。满值进度条值
                self.progress.update()
                time.sleep(0.0001)  # 微小的延迟，实现平滑更新

            # 表格页合并完成后，重新启用“发电数据处理”按钮 。条件为假时，执行下面代码
            if not self.excel_process_button_disabled:
                self.excel_process_button.config(state=tk.NORMAL)

            # 处理待机当天燃料的消耗（，重新启用“处理待机当天燃料的消耗（”按钮 。条件为假时，执行下面代码
            if not self.no_process_button_disabled:
                self.no_process_button.config(state=tk.NORMAL)

            # 停止进度条动画
            self.progress.stop()

    # ==========================================================#

    ##############    开启线程    ##################

    # 处理待机燃料消耗

    # 开启多线程主，调用函数，调用process_data
    def background_no_process_data(self):

        # 当按下“发电数据”时，禁用“合并数据”按钮
        self.process_button.config(state=tk.DISABLED)
        self.process_button_disabled = False

        # 当按下“发电数据”时，禁用“合并数据”按钮
        self.excel_process_button.config(state=tk.DISABLED)
        self.excel_process_button_disabled = False

        threa = Thread(target=self.no_process_data)
        threa.start()

    ##############    关闭线程    ##################

    def no_process_data(self):

        self.progress.start()
        self.progress['value'] += 0  # 逐步增加进度条值

        S_RemFuelIn_value = []

        start_datatime = []
        end_datatime = []
        start_S_RemFuelIn = []
        end_S_RemFuelIn = []
        start_Topgen = []
        end_Topgen = []
        start_S_RemFuelOut = []
        end_S_RemFuelOut = []
        Stwtims = []
        Fuel_consumption = None

        ###################   计算待机燃料消耗    #############

        No_S_RemFuelIn_value = []  # 不发电时，储存内置液位的值
        # NO_differences = []
        NO_positive_differences = []
        NO_Once_S_RemFuelIn = []
        New_MSW = []
        all_Sum_S_RemFuelIn = []
        Timer_RemFuelIn = []

        start_S_RemFuelIn = []
        end_S_RemFuelIn = []

        start_S_RemFuelOut = []
        end_S_RemFuelOut = []
        No_S_RemFuelOut_value = []

        No_LiqlelL = []
        No_LiqlelM = []

        start_No_LiqlelL = []
        end_No_LiqlelL = []

        start_No_LiqlelM = []
        end_No_LiqlelM = []

        NO_DateTime = []
        One_DateTime = []

        No_HGHpre = []
        No_HGHpre_Count = []
        No_HGHpre_SumCount = []

        q = []
        df = []
        df_list = []

        NO_DateTime = []
        One_DateTime = []

        No_HGHpre_time_list = []
        No_HGHpre_time_average = []  # 平均产氢时间
        remark = []  # 备注
        New_StaV = []  # 电堆电压列表

        New_Stapow = []  # 电堆功率列表

        ###################   计算待机燃料消耗    #############

        adress1 = self.file_path  # 读取文件路径。将选择的文件路径赋值给adress1变量
        adress3 = self.save_path  # 保存文件路径

        if not adress1 or not adress3:  # 假设 self.file_path 和 self.save_path 分别表示文件路径和保存路径
            messagebox.showerror("错误", "请选择文件路径和文件保存路径")

            for _ in range(105):
                self.progress['value'] += 1  # 如果，异常。满值进度条值
                self.progress.update()
                time.sleep(0.0001)  # 微小的延迟，实现平滑更新

            # 表格页合并完成后，重新启用“发电数据处理”按钮 。条件为假时，执行下面代码
            if not self.process_button_disabled:
                self.process_button.config(state=tk.NORMAL)

            # 表格页合并完成后，重新启用“发电数据处理”按钮 。条件为假时，执行下面代码
            if not self.excel_process_button_disabled:
                self.excel_process_button.config(state=tk.NORMAL)

            # 停止进度条动画
            self.progress.stop()
        try:
            if os.path.exists(adress1):  # 检查文件（文件名，文件路径是对得上）是否存在，不存在则结束程序
                try:

                    xl = pd.ExcelFile(adress1)  # 使用 pd.ExcelFile() 方法打开 Excel 文件
                    # df = pd.DataFrame()  # 创建一个空的数据框
                    for sheet_name in xl.sheet_names:  # 遍历文件中的所有 sheet
                        one_sheet = xl.parse(sheet_name)  # 读取当前 sheet 的数据
                        df.append(one_sheet)  # 将读取的数据合并到 all_data 中
                    # 使用 pd.concat() 方法将所有数据框连接成一个
                    df = pd.concat(df, ignore_index=True)
                    # 现在 all_data 包含了所有 sheet 的数据

                    df.fillna(0, inplace=True)

                    # 选择要读取的列名
                    MSw = 'MSw'  # 开关状态
                    DateTime = 'DateTime'  # 时间
                    S_RemFuelIn = 'S_RemFuelIn'  # 内置水箱液位
                    S_RemFuelOut = 'S_RemFuelOut'  # 外置水箱液位
                    HGHpre = 'HGHpre'  # 氢气压力

                    # # 获取 'MSw' 列中含有 'nan' 的行的索引
                    # index_nan = df[df['MSw'].isna()].index
                    # 从 'MSw' 列中删除含有 'nan' 的行的索引
                    New_MSW = df['MSw'].tolist()
                    max_index = df.index.max()
                    print(f'索引最大值：{max_index}')
                    # New_MSW=df.dropna(subset=['MSw'],how='any',inplace=True)
                    # print(New_MSW)

                    New_StaV = df['StaV'].tolist()
                    New_Stapow = df['Stapow'].tolist()

                    LiqlelL = 'LiqlelL'  # 外置液位（mm）
                    LiqlelM = 'LiqlelM'  # 内置液位（mm）

                    #   如果电压小于85，则跳过当天计算
                    if any(df['StaV'] >= 0):
                        # second_row = df.iloc[1]  # 这行代码将DataFrame中的第二行数据存储在变量second_row中，以便后续对第二行数据进行操作和分析
                        # last_row = df.iloc[-1]  # 这行代码将DataFrame中的最后一行数据存储在变量last_row中，以便后续对最后一行数据进行操作和分析

                        NO_DateTime = df['DateTime'].tolist()
                        # 分割时间为{年-月-日  ， 时-分-秒}
                        date_only = NO_DateTime[1].split(" ")
                        # 日期 ：年-月-日

                        print(f'\n ————————————————  {date_only[0]}   一天计算开始    ————————————————    \n')

                        # 获取 'MSw' 列的所有数据，并存储到列表 New_MSW 中
                        # filtered_values = [value for value in New_MSW if not math.isnan(value)]
                        # 使用 all() 函数检查 'MSw' 列中的所有值是否都为 False
                        # 使用列表推导式来排除 None 值
                        # filtered_MSW = [value for value in New_MSW if value is not None]
                        # 创建了一个新的迭代器，它只包含New_MSW中不是None的元素。然后，all()
                        # 函数检查这些经过过滤的元素是否都是False。

                        # 使用 all() 函数检查 'MSw' 列中的所有值是否都为 False
                        # 如果MSW=FALSE，不发电时，储存发电时间段内某列的数据，或者 电堆电压StaV全部等于0时，
                        if (all(value == False for value in New_MSW)
                                or all(value == 0 for value in New_StaV)
                                or all(value == 0 for value in New_Stapow)):  # 如果MSW=FALSE，不发电时，储存发电时间段内某列的数据
                            for index, row in df.iterrows():  # 这段代码会遍历 DataFrame df 中的每一行数据。
                                No_S_RemFuelIn_value.append(
                                    round(row[S_RemFuelIn], 1))  # 不发电时，储存 内置水箱剩余燃料 的值到列表 S_RemFuelIn_value
                                No_S_RemFuelOut_value.append(round(row[S_RemFuelOut], 1))
                                No_LiqlelL.append(round(row[LiqlelL], 1))
                                No_LiqlelM.append(round(row[LiqlelM], 1))
                                No_HGHpre.append(round(row[HGHpre], 1))

                            One_DateTime.append(date_only[0])
                            # print(f'时间：{NO_DateTime}')

                            # 内置液位(L)
                            start_S_RemFuelIn.append(No_S_RemFuelIn_value[1])
                            end_S_RemFuelIn.append(No_S_RemFuelIn_value[-1])
                            # 外置液位(L)
                            start_S_RemFuelOut.append(No_S_RemFuelOut_value[1])
                            end_S_RemFuelOut.append(No_S_RemFuelOut_value[-1])
                            # 外置液位(mm)
                            start_No_LiqlelL.append(No_LiqlelL[1])
                            end_No_LiqlelL.append(No_LiqlelL[-1])
                            # 内置液位(mm)
                            start_No_LiqlelM.append(No_LiqlelM[1])
                            end_No_LiqlelM.append(No_LiqlelM[-1])

                            HGHpre_time = 0
                            last_HGHpre_time = 0
                            current_HGHpre_time = 0

                            # 计算产氢次数
                            # 遍历列表中的元素
                            i = 0
                            while i < len(No_HGHpre) - 1:
                                differences = No_HGHpre[i] - No_HGHpre[i + 1]
                                if differences < -1.5 and No_HGHpre[i + 1] > 22.5:
                                    No_HGHpre_Count.append(No_HGHpre[i + 1])
                                    index_time = i + 1

                                    # 计算产氢时间
                                    current_HGHpre_time = datetime.strptime(NO_DateTime[index_time],
                                                                            '%Y-%m-%d %H:%M:%S')

                                    if current_HGHpre_time and last_HGHpre_time:
                                        HGHpre_time = round(
                                            (current_HGHpre_time - last_HGHpre_time).total_seconds() / 60, 2)
                                        print(f'当前时间点：{current_HGHpre_time}  ====  上个时间点：{last_HGHpre_time}')
                                    last_HGHpre_time = current_HGHpre_time
                                    if HGHpre_time:
                                        # 储存平均产氢时间差的值到列表No_HGHpre_time_list
                                        No_HGHpre_time_list.append(HGHpre_time)
                                        print(f"时间差：{HGHpre_time} 分钟")
                                    if max_index > 15000:
                                        i += 3000
                                    elif max_index > 10000:
                                        i += 1250
                                    elif max_index > 7500:
                                        i += 850
                                    elif max_index > 5000:
                                        i += 500
                                    elif max_index > 3000:
                                        i += 280
                                    else:
                                        # 如果条件满足，跳过接下来的200个元素
                                        i += 100  # 增加i的值，确保跳过200个元素

                                else:
                                    # 如果条件不满足，正常递增i
                                    i += 1  # 正常递增i

                            #     q.append(i)
                            # print(f"循环列表：======{q}")
                            if len(No_HGHpre_time_list) > 2:
                                average = round(sum(No_HGHpre_time_list) / len(No_HGHpre_time_list), 2)
                            else:
                                average = 0
                            No_HGHpre_time_average.append(average)
                            # print(f'平均')
                            print(f'平均产氢时间：{average}')
                            print(f"计算产气次数 ：{len(No_HGHpre_Count)}")

                            No_HGHpre_SumCount.append(len(No_HGHpre_Count))
                            # print(f'时间  列表：{One_DateTime}')
                            No_HGHpre_Count.clear()

                            print(f"时间 ：{date_only[0]}")
                            print(f'开始时内置液位(mm)：{start_No_LiqlelL[-1]}')
                            print(f'结束时内置液位(mm)：{end_No_LiqlelL[-1]}')
                            print(f'开始时外置液位(mm)：{start_No_LiqlelM[-1]}')
                            print(f'结束时外置液位(mm)：{end_No_LiqlelM[-1]}')

                            print(f'开始时内置液位(L)：{start_S_RemFuelIn[-1]}')
                            print(f'结束时内置液位(L)：{end_S_RemFuelIn[-1]}')
                            print(f'开始时外置液位(L)：{start_S_RemFuelOut[-1]}')
                            print(f'结束时外置液位(L)：{end_S_RemFuelOut[-1]}')

                            Max_Msw = max(No_S_RemFuelIn_value)
                            print(f"最大值:{Max_Msw}")
                            # print(f'燃料值（L）:{No_S_RemFuelIn_value}')

                            Max_Msw_mm = max(No_LiqlelM)
                            print(f"最大值*（mm）:{Max_Msw_mm}")

                            Min_Msw_mm = min(No_LiqlelM)
                            print(f"最小值*（mm）:{Min_Msw_mm}")

                            print(f"++++++液位（L）的列表:{start_S_RemFuelIn}")
                            print(f"------液位（MM）的列表:{start_No_LiqlelL}")

                            NO_differences = 0
                            # print('@@@@@@@@@@@@', No_S_RemFuelIn_value[1], No_S_RemFuelOut_value[1])
                            # 如果列表不为空。说明有数据，执行下面计算
                            if No_S_RemFuelIn_value[1] > 0 and No_S_RemFuelOut_value[1] > 0:
                                # print('@@@@@@@@@@@@@@@@@@@@')
                                # 如果一天中有加液，找出最大值去减第一项，大于1。说明当天有加液
                                if (Max_Msw - No_S_RemFuelIn_value[1]) > 1:
                                    first_RemFuelIn = No_S_RemFuelIn_value[1] - 15
                                    second_RemFuelIn = Max_Msw - No_S_RemFuelIn_value[-1]
                                    NO_differences = round(first_RemFuelIn + second_RemFuelIn, 2)
                                    # print(f'燃料值的差====（L）:{NO_differences}')
                                    # print(f"最大值-第一个:{Max_Msw - No_S_RemFuelIn_value[1]}")
                                    # 如果计算出来的NO_differences液位消耗值小于0，则等于0
                                    if NO_differences <= 0:
                                        NO_differences = 0
                                    # NO_Once_RemFuelIn = round(sum(NO_differences), 2)
                                    NO_Once_S_RemFuelIn.append(NO_differences)
                                    print(f'当天有加液 + 不发电消耗燃料（L）:{NO_differences} 内部')

                                else:

                                    NO_differences = round(No_S_RemFuelIn_value[1] - No_S_RemFuelIn_value[-1], 2)
                                    # 如果计算出来的NO_differences液位消耗值小于0，则等于0
                                    if NO_differences <= 0:
                                        NO_differences = 0
                                    NO_Once_S_RemFuelIn.append(NO_differences)
                                    # print('当天没有加液  ， 计算 L')
                                    print(f'当天没有加液 + 不发电消耗燃料（L）:{NO_differences}  ')
                            # 如果列表为空。说明没有数据，执行下面计算
                            else:
                                # 如果一天中有加液，找出最大值去减第一项，大于1。说明当天有加液
                                if (Max_Msw_mm - No_LiqlelM[1]) > 50:
                                    first_RemFuelIn = No_LiqlelM[1] - Min_Msw_mm
                                    second_RemFuelIn = Max_Msw - No_LiqlelM[-1]
                                    NO_differences = round(first_RemFuelIn + second_RemFuelIn, 2)
                                    # print(f'燃料值的差====（L）:{NO_differences}')
                                    # print(f"最大值-第一个:{Max_Msw - No_S_RemFuelIn_value[1]}")
                                    # 如果计算出来的NO_differences液位消耗值小于0，则等于0
                                    if NO_differences <= 0:
                                        NO_differences = 0
                                    # NO_Once_RemFuelIn = round(sum(NO_differences), 2)
                                    NO_Once_S_RemFuelIn.append(NO_differences)
                                    print(f'当天有加液 + 不发电消耗燃料（mm）:{NO_differences} 内部')

                                else:

                                    NO_differences = round(No_LiqlelM[1] - No_LiqlelM[-1], 2)
                                    # 如果计算出来的NO_differences液位消耗值小于0，则等于0
                                    if NO_differences <= 0:
                                        NO_differences = 0
                                    NO_Once_S_RemFuelIn.append(NO_differences)
                                    # print('当天没有加液  ， 计算 mm')
                                    print(f'当天没有加液 + 不发电消耗燃料（mm）:{NO_differences}  ')

                            # 将待机时的 a1时间 添加到 Timer_RemFuelIn 数组里面。里面只包含待机时间数据
                            Timer_RemFuelIn.append(date_only[0])
                            # 将待机时的 NO_Once_S_RemFuelIn 液位消耗 求出总和
                            Sum_S_RemFuelIn = sum(NO_Once_S_RemFuelIn)
                            all_Sum_S_RemFuelIn.append(Sum_S_RemFuelIn)
                            remark.append(0)
                            print(f'\n===========   {date_only[0]} 当天待机燃料消耗   ==========\n')
                        else:

                            print(filtered_values)

                            b1 = f'当天有发电，不计算待机燃料消耗'
                            o1 = 0
                            start_No_LiqlelL.append(o1)
                            end_No_LiqlelL.append(o1)
                            start_No_LiqlelM.append(o1)
                            end_No_LiqlelM.append(o1)
                            start_S_RemFuelOut.append(o1)
                            end_S_RemFuelOut.append(o1)
                            start_S_RemFuelIn.append(o1)
                            end_S_RemFuelIn.append(o1)
                            No_HGHpre_SumCount.append(o1)
                            all_Sum_S_RemFuelIn.append(o1)
                            One_DateTime.append(date_only[0])
                            No_HGHpre_time_average.append(o1)
                            remark.append(b1)

                            print(f'\n===========   {date_only[0]} 当天有发电，不计算燃料消耗   ==========\n')

                        # print(f"总燃料消耗(L)：{Sum_S_RemFuelIn}")

                        NO_Once_S_RemFuelIn.clear()
                        No_S_RemFuelIn_value.clear()
                        No_S_RemFuelOut_value.clear()
                        No_LiqlelL.clear()
                        No_LiqlelM.clear()
                        NO_DateTime.clear()
                        No_HGHpre.clear()
                        q.clear()
                        df_list.clear()
                        No_HGHpre_time_list.clear()
                        # 在控制台上打印，显示每列的长度(元素个数) ，如果长度(元素个数)不一样，会报错“输出的列长不一样”

                        print(f"\n时间 长度：{len(One_DateTime)}")
                        print(f"消耗燃料 长度：{len(all_Sum_S_RemFuelIn)}")

                        print(f"内置-开始时液位(L) 长度：{len(start_S_RemFuelIn)}")
                        print(f"内置-结束时液位(L) 长度：{len(end_S_RemFuelIn)}")
                        print(f"外置-开始时液位(L） 长度：{len(start_S_RemFuelOut)}")
                        print(f"外置-结束时液位(L) 长度：{len(end_S_RemFuelOut)}")
                        print(f"外置-结束时液位(MM) 长度：{len(start_No_LiqlelL)}")
                        print(f"外置-结束时液位(MM) 长度：{len(end_No_LiqlelL)}")
                        print(f"内置-结束时液位(MM) 长度：{len(start_No_LiqlelM)}")
                        print(f"内置-结束时液位(MM) 长度：{len(end_No_LiqlelM)}\n")
                        print(f"产氢次数 长度：{len(No_HGHpre_SumCount)}")

                        print(f"平均产氢时间 长度：{len(No_HGHpre_time_average)}")
                        print(f"备注 长度：{len(remark)}")

                        print(f'\n++++++++++++++  {date_only[0]} 一天的计算结束   ++++++++++++++++++++++++\n')

                        # 储存 a1时间点 到 Timer_RemFuelIn列表 里面，用于在excel表格打印

                    else:
                        b1 = f' 当天没有数据，下载数据为空 ！！！'
                        o1 = 0

                        One_DateTime.append(One_DateTime[-1])

                        start_No_LiqlelL.append(o1)
                        end_No_LiqlelL.append(o1)
                        start_No_LiqlelM.append(o1)
                        end_No_LiqlelM.append(o1)
                        start_S_RemFuelOut.append(o1)
                        end_S_RemFuelOut.append(o1)
                        start_S_RemFuelIn.append(o1)
                        end_S_RemFuelIn.append(o1)
                        all_Sum_S_RemFuelIn.append(o1)
                        No_HGHpre_SumCount.append(o1)
                        No_HGHpre_time_average.append(o1)

                        remark.append(b1)
                        print(
                            f'\n++++++++++++++   {One_DateTime[-1]}    当天没有数据，下载数据为空 ！！！    ++++++++++++++++++++++++\n')

                except FileNotFoundError:
                    print(f"文件 {adress1} 不存在，已跳过")
            else:
                print(f"文件 {adress1} 不存在，已跳过")

            A_all_Sum_S_RemFuelIn = sum(all_Sum_S_RemFuelIn)

            print(f"总燃料消耗(L)：{A_all_Sum_S_RemFuelIn}\n")

            # print(f"时间：{Timer_RemFuelIn}\n")

            # 在控制台上打印，显示每列的长度(元素个数) ，如果长度(元素个数)不一样，会报错“输出的列长不一样”
            print(f"时间 长度：{len(One_DateTime)}")
            # print(f"时间 长度：{len(Timer_RemFuelIn)}")
            print(f"消耗燃料 长度：{len(all_Sum_S_RemFuelIn)}")

            print(f"内置-开始时液位(L) 长度：{len(start_S_RemFuelIn)}")
            print(f"内置-结束时液位(L) 长度：{len(end_S_RemFuelIn)}")
            print(f"外置-开始时液位(L） 长度：{len(start_S_RemFuelOut)}")
            print(f"外置-结束时液位(L) 长度：{len(end_S_RemFuelOut)}")
            print(f"外置-结束时液位(MM) 长度：{len(start_No_LiqlelL)}")
            print(f"外置-结束时液位(MM) 长度：{len(end_No_LiqlelL)}")
            print(f"内置-结束时液位(MM) 长度：{len(start_No_LiqlelM)}")
            print(f"内置-结束时液位(MM) 长度：{len(end_No_LiqlelM)}")
            print(f"产氢次数 长度：{len(No_HGHpre_SumCount)}")

            print(f"平均产氢时间 长度：{len(No_HGHpre_time_average)}")
            print(f"备注 长度：{len(remark)}")
            # 将新的DataFrame保存到新的Excel文件中
            #
            # print(f'start_S_RemFuelOut======{start_S_RemFuelOut[0]}')
            # print(f'end_S_RemFuelOut======{end_S_RemFuelOut[0]}')

            if start_S_RemFuelOut[0] > 0 and end_S_RemFuelOut[0] > 0:
                # 将新的DataFrame保存到新的Excel文件中
                new_df = pd.DataFrame(
                    {
                        '时间': One_DateTime,
                        # '时间': Timer_RemFuelIn,
                        '开始外置水箱剩余燃料(mm)': start_No_LiqlelL,
                        '结束外置水箱剩余燃料(mm)': end_No_LiqlelL,
                        '开始内置水箱剩余燃料(mm)': start_No_LiqlelM,
                        '结束内置水箱剩余燃料(mm)': end_No_LiqlelM,
                        '开始外置水箱剩余燃料(L)': start_S_RemFuelOut,
                        '结束外置水箱剩余燃料(L)': end_S_RemFuelOut,
                        '开始内置水箱剩余燃料(L)': start_S_RemFuelIn,
                        '结束内置水箱剩余燃料(L)': end_S_RemFuelIn,

                        '待机消耗燃料(L)': all_Sum_S_RemFuelIn,
                        '产氢计数（次）': No_HGHpre_SumCount,
                        '平均产氢时间（min）': No_HGHpre_time_average,
                        '备注': remark,

                    })

            else:

                # 将新的DataFrame保存到新的Excel文件中
                new_df = pd.DataFrame(
                    {
                        '时间': One_DateTime,
                        # '时间': Timer_RemFuelIn,
                        '开始外置水箱剩余燃料(mm)': start_No_LiqlelL,
                        '结束外置水箱剩余燃料(mm)': end_No_LiqlelL,
                        '开始内置水箱剩余燃料(mm)': start_No_LiqlelM,
                        '结束内置水箱剩余燃料(mm)': end_No_LiqlelM,
                        '开始外置水箱剩余燃料(L)': start_S_RemFuelOut,
                        '结束外置水箱剩余燃料(L)': end_S_RemFuelOut,
                        '开始内置水箱剩余燃料(L)': start_S_RemFuelIn,
                        '结束内置水箱剩余燃料(L)': end_S_RemFuelIn,

                        '待机消耗燃料(mm)': all_Sum_S_RemFuelIn,
                        '产氢计数（次）': No_HGHpre_SumCount,
                        '平均产氢时间（min）': No_HGHpre_time_average,
                        '备注': remark,

                    })

            file_path = adress3
            new_df.to_excel(file_path, index=False)
            # 打开现有的Excel文件
            workbook = openpyxl.load_workbook(file_path)
            # 选择第一个工作表
            sheet = workbook.active
            # 设置第一行的行高
            sheet.row_dimensions[1].height = 50
            # 设置第一列和第二列的宽度为 25
            sheet.column_dimensions['A'].width = 21  # 第一列
            # sheet.column_dimensions['B'].width = 21  # 第二列
            # 设置其余列的宽度为 10
            for col in sheet.columns:
                if col[0].column_letter not in ['A']:
                    sheet.column_dimensions[col[0].column_letter].width = 15
            # 遍历第一行的所有单元格，并为每个单元格对象同时设置自动换行、水平居中和垂直居中。
            for cell in sheet[1]:
                cell_obj = cell
                cell_obj.alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal='center',
                                                               vertical='center')

            workbook.save(file_path)

            # 表格页合并完成后，重新启用“发电数据处理”按钮 。条件为假时，执行下面代码
            if not self.process_button_disabled:
                self.process_button.config(state=tk.NORMAL)

            # 表格页合并完成后，重新启用“发电数据处理”按钮 。条件为假时，执行下面代码
            if not self.excel_process_button_disabled:
                self.excel_process_button.config(state=tk.NORMAL)

            # 完成所有操作后更新进度条到100%
            for _ in range(10):
                self.progress['value'] += 1  # 逐步增加进度条值
                self.progress.update()
                time.sleep(0.05)  # 微小的延迟，实现平滑更新
            # 停止进度条动画
            self.progress.stop()

            print(f"\n文件保存成功 ！! ! ")
            print(f"文件保存路径 ：{file_path}")
            self.show_save_success_message(adress3)

        # system_state.clear()

        except ValueError:

            for _ in range(105):
                self.progress['value'] += 1  # 如果，异常。满值进度条值
                self.progress.update()
                time.sleep(0.0001)  # 微小的延迟，实现平滑更新

            # 表格页合并完成后，重新启用“发电数据处理”按钮 。条件为假时，执行下面代码
            if not self.process_button_disabled:
                self.process_button.config(state=tk.NORMAL)

            # 表格页合并完成后，重新启用“发电数据处理”按钮 。条件为假时，执行下面代码
            if not self.excel_process_button_disabled:
                self.excel_process_button.config(state=tk.NORMAL)

            # 停止进度条动画
            self.progress.stop()
            self.show_read_error(adress1)


if __name__ == "__main__":
    root = tk.Tk()
    app = DataProcessingApp(root)

    root.mainloop()
